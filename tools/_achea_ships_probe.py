#!/usr/bin/env python3
"""Read-only: actor26 vs water, special id census, PL8 boat sheets, EXE strings."""

from __future__ import annotations

import struct
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

SAV = Path(r"C:\Users\Felip\OneDrive\Games\Caesar2\Achea.sav\ACHEA23.SAV")
GAME = Path(r"C:\Users\Felip\OneDrive\Games\Caesar2")
XLSX = Path(r"C:\Users\Felip\caesar2-re\findings\Achea_province.xlsx")
BIN = Path(r"C:\Users\Felip\caesar2-re\ghidra_work\c2_x.bin")

CHUNK7_OFF = 16
CHUNK14_OFF = 178395
MAP_W = 60
REC = 8
ACTOR_STRIDE = 175
ID_SPECIAL = 0x7D


def rec_at(blob, x, y):
    i = (y * MAP_W + x) * REC
    return blob[i : i + REC]


def main():
    sav = SAV.read_bytes()
    blob = sav[CHUNK14_OFF : CHUNK14_OFF + 28800]
    ids = [[rec_at(blob, x, y)[0] for x in range(60)] for y in range(60)]

    wb = load_workbook(XLSX, data_only=True)
    ws = wb["mapa"]
    labels = [[(ws.cell(y + 2, x + 2).value or "") for x in range(60)] for y in range(60)]

    print("=== special id census ===")
    by_id = defaultdict(list)
    for y in range(60):
        for x in range(60):
            if ids[y][x] >= ID_SPECIAL:
                by_id[ids[y][x]].append((x, y, str(labels[y][x]).strip()))
    for tid in sorted(by_id):
        cells = by_id[tid]
        names = Counter(c[2] or "(blank)" for c in cells)
        xs = [c[0] for c in cells]
        ys = [c[1] for c in cells]
        w = max(xs) - min(xs) + 1
        h = max(ys) - min(ys) + 1
        print(
            f"  0x{tid:02X} n={len(cells):3} bbox=({min(xs)},{min(ys)})-({max(xs)},{max(ys)}) "
            f"{w}x{h} names={dict(names)}"
        )

    print("=== water-ish tiles near sea-lane portals ===")
    portals = [(0, 44, "Campania 0x9F"), (40, 59, "Creta 0x9E"), (59, 10, "Trade 0x9D")]
    for px, py, name in portals:
        print(f"  {name} tile=0x{ids[py][px]:02X}")
        for dy in range(-2, 3):
            row = []
            for dx in range(-2, 3):
                x, y = px + dx, py + dy
                if 0 <= x < 60 and 0 <= y < 60:
                    row.append(f"{ids[y][x]:02X}")
                else:
                    row.append("--")
            print("   ", " ".join(row))

    print("=== type6 neighborhood (5x5 byte0) ===")
    a26 = sav[CHUNK7_OFF : CHUNK7_OFF + 4550]
    for i in range(26):
        rec = a26[i * ACTOR_STRIDE : (i + 1) * ACTOR_STRIDE]
        if rec[0] == 0:
            continue
        x, y = rec[6], rec[7]
        typ = rec[4]
        print(f"  slot {i} type={typ} ({x},{y}) dest=({rec[0xE]},{rec[0xF]}) state={rec[0x12]}")
        for dy in range(-3, 4):
            row = []
            for dx in range(-3, 4):
                xx, yy = x + dx, y + dy
                if 0 <= xx < 60 and 0 <= yy < 60:
                    mark = "*" if dx == 0 and dy == 0 else " "
                    row.append(f"{ids[yy][xx]:02X}{mark}")
                else:
                    row.append(" -- ")
            print("   ", " ".join(row))

    print("=== water id histogram (byte0 < 0x18) ===")
    water = Counter()
    for y in range(60):
        for x in range(60):
            if ids[y][x] < 0x18:
                water[ids[y][x]] += 1
    print(" ", {hex(k): v for k, v in sorted(water.items())}, "total", sum(water.values()))

    print("=== PL8 headers boat/unit/prov ===")
    for name in [
        "PUNBOAT.PL8",
        "ASHIPYA.PL8",
        "MY_STDS.PL8",
        "MY_STDS2.PL8",
        "MY_STDS3.PL8",
        "PROVFIXT.PL8",
        "PROV1.PL8",
        "INT_PROV.PL8",
        "LTLMEN1B.PL8",
    ]:
        p = GAME / name
        if not p.exists():
            # case
            hits = list(GAME.glob(name))
            p = hits[0] if hits else None
        if not p:
            print(f"  {name} MISSING")
            continue
        data = p.read_bytes()
        flags, n = struct.unpack_from("<HH", data, 0)
        # sample sprite sizes around 0x4E
        print(f"  {p.name} flags=0x{flags:04X} n={n} size={len(data)}")
        for idx in (0, 1, 6, 8, 0x12, 0x1A, 0x36, 0x4E, 0x4F, 0x51, 0x54, n - 1):
            if idx < 0 or idx >= n:
                continue
            off = 8 + 16 * idx
            w, h, doff, sx, sy, t, extra = struct.unpack_from("<HHIHHBB", data, off)
            print(f"    spr {idx:3} {w}x{h} type={t} extra={extra} xy=({sx},{sy})")

    print("=== EXE / bin strings ship/boat/pun/sea ===")
    blobs = []
    exe = GAME / "PS.EXE"
    if exe.exists():
        blobs.append(("PS.EXE", exe.read_bytes()))
    if BIN.exists():
        blobs.append(("c2_x.bin", BIN.read_bytes()))
    keys = (
        b"punboat",
        b"PUNBOAT",
        b"aship",
        b"ASHIP",
        b"ship5",
        b"SHIP5",
        b"Sea Lane",
        b"sea lane",
        b"SEALANE",
        b"boat",
        b"Boat",
        b"ship",
    )
    for label, data in blobs:
        print(f"  -- {label} {len(data)} --")
        for k in keys:
            idx = 0
            found = 0
            while True:
                i = data.find(k, idx)
                if i < 0:
                    break
                found += 1
                snippet = data[max(0, i - 8) : i + 24]
                snippet = bytes(b if 32 <= b < 127 else 46 for b in snippet)
                print(f"    {k!s} @ {i:#x} {snippet!s}")
                idx = i + 1
                if found >= 8:
                    print("    ...")
                    break
            if not found:
                print(f"    {k!s} not found")

    help_eng = GAME / "HELP.ENG"
    c2eng = GAME / "C2.ENG"
    for p in (help_eng, c2eng):
        if not p.exists():
            print(f"  {p.name} missing")
            continue
        data = p.read_bytes()
        print(f"=== {p.name} ship/sea/lane/boat ===")
        text = data.decode("latin-1", "replace")
        for needle in ("Sea", "ship", "Ship", "boat", "Boat", "lane", "Lane", "Port", "navy"):
            if needle.lower() in text.lower():
                # print nearby
                pass
        low = text.lower()
        for needle in ("sea lane", "shipyard", "ship", "boat", "port", "trading"):
            i = 0
            n = 0
            while n < 5:
                j = low.find(needle, i)
                if j < 0:
                    break
                frag = text[max(0, j - 20) : j + 40].replace("\n", " ").replace("\r", " ")
                print(f"    {needle!r} ... {frag!r}")
                i = j + len(needle)
                n += 1


if __name__ == "__main__":
    main()
