#!/usr/bin/env python3
"""Dump housing ids, tile[+15] land-value, C2MODEL occupancy, EXE evolve table.

Reads SAV / C2MODEL / PS.EXE in place. Does not copy binaries into git.
"""

from __future__ import annotations

import csv
import statistics
import struct
import sys
from collections import Counter, defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / "tools"))

from app.city_map import ID_HOUSING_HI, ID_HOUSING_LO, MAP_H, MAP_W, load_city_from_sav  # noqa: E402
from dump_c2model import HOUSE_GRADES  # noqa: E402
from ps_le import DEFAULT_EXE, load_ps, map_image  # noqa: E402

GAME = Path(r"C:\Users\Felip\OneDrive\Games\Caesar2")
C2MODEL = GAME / "C2MODEL.DAT"

# FAQ / caesar2.com building directory (names only; LV column on that page).
# 32 grades ↔ ids 0x82..0xA1 (city_buildings_evolve_row: grade = id + 0x7E as i8).
DIR_REQUIRED_LV = [
    0, 2, 4, 6, 8, 10, 12, 14, 16, 18, 20, 22, 24, 26, 28, 30,
    32, 34, 36, 38, 40, 42, 44, 46, 48, 50, 52, 54, 58, 60, 62, 64,
]

SAV_CANDIDATES = [
    GAME / "Achea.sav" / "ACHEA23.SAV",
    GAME / "ACHEA23.SAV",
    ROOT / "findings" / "D.SAV",
    GAME / "D.SAV",
    GAME / "20230610.SAV",
    GAME / "20230610.sav" / "20230610.SAV",
    GAME / "FELIPE01.SAV",
    GAME / "FELIPE02.SAV",
    GAME / "LASTYEAR.SAV",
    ROOT / "findings" / "20230610.SAV",
]


def i8(b: int) -> int:
    return b - 256 if b >= 128 else b


def signed_stats(values: list[int]) -> dict[str, float | int]:
    if not values:
        return {"n": 0, "min": "", "med": "", "mean": "", "max": "", "uniq": ""}
    s = sorted(values)
    n = len(s)
    if n % 2:
        med = float(s[n // 2])
    else:
        med = (s[n // 2 - 1] + s[n // 2]) / 2.0
    return {
        "n": n,
        "min": min(s),
        "med": med,
        "mean": round(statistics.mean(s), 2),
        "max": max(s),
        "uniq": len(set(s)),
    }


def find_savs() -> list[Path]:
    found: list[Path] = []
    seen: set[str] = set()
    for p in SAV_CANDIDATES:
        if p.is_file():
            key = p.resolve().as_posix().lower()
            if key not in seen:
                seen.add(key)
                found.append(p)
    extra_roots = [GAME, GAME / "Achea.sav", ROOT / "findings"]
    for root in extra_roots:
        if not root.is_dir():
            continue
        for p in root.glob("*.SAV"):
            key = p.resolve().as_posix().lower()
            if key not in seen:
                seen.add(key)
                found.append(p)
        for p in root.glob("*.sav"):
            if p.suffix.lower() == ".sav" and p.is_file():
                key = p.resolve().as_posix().lower()
                if key not in seen:
                    seen.add(key)
                    found.append(p)
    return found


def housing_from_sav(path: Path) -> dict:
    city = load_city_from_sav(path, game=GAME if (GAME / "PS.EXE").is_file() else path.parent)
    by_id: dict[int, list[tuple[int, int, int, int]]] = defaultdict(list)
    # (x, y, lv15, origin)
    for y in range(MAP_H):
        for x in range(MAP_W):
            t = city.tile(x, y)
            hid = t.terrain_id
            if not (ID_HOUSING_LO <= hid <= ID_HOUSING_HI):
                continue
            lv = i8(t.industry)  # tile[+15]
            origin = (t.spawn_packed & 0xF) == 0
            by_id[hid].append((x, y, lv, origin))
    return {"path": path, "by_id": by_id}


def dump_exe_tables() -> dict:
    mapped = map_image(load_ps(DEFAULT_EXE), apply_fixups=True)
    ev = bytes(mapped.image[mapped.va_to_off(0x96235) : mapped.va_to_off(0x96235) + 64])
    size_lut = bytes(mapped.image[mapped.va_to_off(0x94F40) : mapped.va_to_off(0x94F40) + 32 * 4])
    bonus = bytes(mapped.image[mapped.va_to_off(0x962FD) : mapped.va_to_off(0x962FD) + 32 * 8])
    pairs = []
    for i in range(32):
        lo = i8(ev[i * 2])
        hi = i8(ev[i * 2 + 1])
        sz = struct.unpack_from("<i", size_lut, i * 4)[0]
        bns, rad = struct.unpack_from("<ii", bonus, i * 8)
        pairs.append({"grade": i, "min": lo, "max": hi, "size": sz, "bonus": bns, "radius": rad})
    return {"pairs": pairs, "raw_ev": ev.hex()}


def dump_c2model() -> dict:
    data = C2MODEL.read_bytes()
    ints = list(struct.unpack_from("<1090i", data))
    occ = ints[215:247]
    tax = ints[247:279]
    land = [(ints[500 + i * 2], ints[501 + i * 2]) for i in range(32)]
    decay = ints[404:436]
    return {"occ": occ, "tax": tax, "land": land, "decay": decay}


def xlsx_house_labels() -> list[str]:
    lines: list[str] = []
    try:
        from openpyxl import load_workbook
    except ImportError:
        return ["openpyxl missing"]
    for name in ("Achea_grid.xlsx", "Achea_grid_v3.xlsx", "Achea_grid_new.xlsx", "20230610_grid.xlsx"):
        p = ROOT / "findings" / name
        if not p.is_file():
            lines.append(f"{name}: missing")
            continue
        wb = load_workbook(p, read_only=True, data_only=True)
        house_labels: Counter[str] = Counter()
        nomes: list[str] = []
        for sheet in wb.worksheets:
            title = (sheet.title or "").lower()
            if title in {"nomes", "names", "legend", "legenda"}:
                for row in sheet.iter_rows(min_row=1, max_col=4, values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    if any("0x8" in c or "0x9" in c or "0xA" in c or "Casa" in c or "Tent" in c or "House" in c for c in cells):
                        nomes.append(" | ".join(cells))
            # scan all cells for Casa/House/Tent near housing
            for row in sheet.iter_rows(values_only=True):
                for c in row:
                    if not isinstance(c, str):
                        continue
                    cl = c.strip()
                    if cl.lower() in {"casa", "house", "tent", "hut"} or cl.startswith("Casa") or cl.startswith("House"):
                        house_labels[cl] += 1
        wb.close()
        top = ", ".join(f"{k}={v}" for k, v in house_labels.most_common(12)) or "(none)"
        lines.append(f"{name} cell labels: {top}")
        for n in nomes[:12]:
            lines.append(f"  nomes: {n}")
    return lines


def main() -> int:
    print("=== xlsx labels ===")
    for line in xlsx_house_labels():
        print(line)

    print("\n=== EXE evolve table 0x96235 (min,max per grade) + 0x94F40 size + 0x962FD bonus ===")
    exe = dump_exe_tables()
    for p in exe["pairs"]:
        print(
            f"  g{p['grade']:02d} id=0x{0x82+p['grade']:02X}  "
            f"ev_min={p['min']:3d} ev_max={p['max']:3d}  size={p['size']}  "
            f"bonus={p['bonus']:3d} r={p['radius']}"
        )
    print("raw", exe["raw_ev"])

    print("\n=== C2MODEL occupancy / land(bonus,r) / decay ===")
    mdl = dump_c2model()
    for i, name in enumerate(HOUSE_GRADES):
        print(
            f"  g{i:02d} {name:22s} occ={mdl['occ'][i]:4d} tax={mdl['tax'][i]:4d} "
            f"land=({mdl['land'][i][0]:3d},{mdl['land'][i][1]}) decay={mdl['decay'][i]:3d}"
        )

    savs = find_savs()
    print("\n=== SAV files ===")
    for p in savs:
        print(f"  {p}  {p.stat().st_size} B")

    reports = []
    for p in savs:
        try:
            reports.append(housing_from_sav(p))
        except Exception as exc:
            print(f"FAIL {p}: {exc}")

    for rep in reports:
        print(f"\n=== {rep['path'].name} housing tiles ===")
        ids = sorted(rep["by_id"])
        print(f"  distinct ids: {len(ids)}  {[hex(i) for i in ids]}")
        for hid in range(ID_HOUSING_LO, ID_HOUSING_HI + 1):
            rows = rep["by_id"].get(hid, [])
            if not rows:
                continue
            all_lv = [r[2] for r in rows]
            orig_lv = [r[2] for r in rows if r[3]]
            st = signed_stats(all_lv)
            sto = signed_stats(orig_lv)
            print(
                f"  0x{hid:02X} tiles={st['n']:3d} orig={sto['n']:3d}  "
                f"+15 all min/med/mean/max={st['min']}/{st['med']}/{st['mean']}/{st['max']} uniq={st['uniq']}  "
                f"orig {sto['min']}/{sto['med']}/{sto['mean']}/{sto['max']} uniq={sto['uniq']}"
            )

    # Write CSV + compact print for Achea
    ache = next((r for r in reports if "ACHEA" in r["path"].name.upper()), reports[0] if reports else None)
    extra = next((r for r in reports if "20230610" in r["path"].name), None)
    fel = next((r for r in reports if "FELIPE01" in r["path"].name.upper()), None)

    out_csv = ROOT / "findings" / "housing_land_value.csv"
    rows_out = []
    header = [
        "id_hex",
        "id_dec",
        "grade",
        "name_faq",
        "achea_tiles",
        "achea_origins",
        "achea_lv15_min",
        "achea_lv15_med",
        "achea_lv15_mean",
        "achea_lv15_max",
        "achea_lv15_uniq",
        "s20230610_tiles",
        "s20230610_lv15_min",
        "s20230610_lv15_med",
        "s20230610_lv15_max",
        "dir_required_lv",
        "exe_ev_min",
        "exe_ev_max",
        "occupancy",
        "c2model_land_bonus",
        "c2model_land_radius",
        "footprint_hint",
    ]
    for i, name in enumerate(HOUSE_GRADES):
        hid = 0x82 + i
        a_rows = ache["by_id"].get(hid, []) if ache else []
        e_rows = extra["by_id"].get(hid, []) if extra else []
        sta = signed_stats([r[2] for r in a_rows])
        ste = signed_stats([r[2] for r in e_rows])
        orig_n = sum(1 for r in a_rows if r[3])
        p = exe["pairs"][i]
        if i < 26:
            fp = "1x1"
        elif i < 30:
            fp = "2x2"
        else:
            fp = "3x3"
        rows_out.append(
            {
                "id_hex": f"0x{hid:02X}",
                "id_dec": hid,
                "grade": i,
                "name_faq": name,
                "achea_tiles": sta["n"],
                "achea_origins": orig_n,
                "achea_lv15_min": sta["min"],
                "achea_lv15_med": sta["med"],
                "achea_lv15_mean": sta["mean"],
                "achea_lv15_max": sta["max"],
                "achea_lv15_uniq": sta["uniq"],
                "s20230610_tiles": ste["n"],
                "s20230610_lv15_min": ste["min"],
                "s20230610_lv15_med": ste["med"],
                "s20230610_lv15_max": ste["max"],
                "dir_required_lv": DIR_REQUIRED_LV[i],
                "exe_ev_min": p["min"],
                "exe_ev_max": p["max"],
                "occupancy": mdl["occ"][i],
                "c2model_land_bonus": mdl["land"][i][0],
                "c2model_land_radius": mdl["land"][i][1],
                "footprint_hint": fp,
            }
        )

    with out_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=header)
        w.writeheader()
        w.writerows(rows_out)
    print(f"\nwrote {out_csv}")

    if extra:
        print(f"\n=== {extra['path'].name} distinct housing ===")
        print(sorted(hex(i) for i in extra["by_id"]))
        print("counts", {hex(i): len(v) for i, v in sorted(extra["by_id"].items())})
    if fel:
        print(f"\n=== {fel['path'].name} distinct housing ===")
        print({hex(i): len(v) for i, v in sorted(fel["by_id"].items())})
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
