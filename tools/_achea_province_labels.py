#!/usr/bin/env python3
"""Read-only: user labels in Achea_province.xlsx vs ACHEA23 chunk 14 + actor26."""

from __future__ import annotations

import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

XLSX = ROOT / "findings" / "Achea_province.xlsx"
SAV = Path(r"C:\Users\Felip\OneDrive\Games\Caesar2\Achea.sav\ACHEA23.SAV")
GAME = Path(r"C:\Users\Felip\OneDrive\Games\Caesar2")

CHUNK7_OFF = 16
CHUNK7_SIZE = 4550
CHUNK8_OFF = 4566
CHUNK8_SIZE = 11658
CHUNK14_OFF = 178395
CHUNK14_SIZE = 28800
MAP_W = MAP_H = 60
REC = 8
ID_SPECIAL = 0x7D
ACTOR_STRIDE = 175
ACTOR_N = 26
WALKER_STRIDE = 58
WALKER_N = 201

SKIP = {
    None,
    "",
    "desconhecido",
    "terreno",
    "terrain",
}


def norm(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return s


def is_placeholder(s: str) -> bool:
    if not s:
        return True
    low = s.lower()
    if low in SKIP:
        return True
    if low.startswith("desconhecido"):
        return True
    return False


def rec_at(blob: bytes, x: int, y: int) -> bytes:
    i = (y * MAP_W + x) * REC
    return blob[i : i + REC]


def main() -> None:
    wb = load_workbook(XLSX, data_only=True)
    print("=== sheets ===")
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"  {name!r} {ws.dimensions} max_row={ws.max_row} max_col={ws.max_column}")

    ws = wb["mapa"]
    # Header row: col 2..61 = x 0..59; row 2..61 = y 0..59
    xs = [ws.cell(1, c).value for c in range(2, 62)]
    ys = [ws.cell(r, 1).value for r in range(2, 62)]
    print("=== mapa axes ===")
    print(f"  header x first/last: {xs[0]} .. {xs[-1]}  unique={len(set(xs))}")
    print(f"  header y first/last: {ys[0]} .. {ys[-1]}  unique={len(set(ys))}")
    if xs != list(range(60)) or ys != list(range(60)):
        print("  WARN: header not 0..59 sequential")

    labels = [[norm(ws.cell(y + 2, x + 2).value) for x in range(60)] for y in range(60)]

    sav = SAV.read_bytes()
    assert len(sav) == 225745, len(sav)
    blob = sav[CHUNK14_OFF : CHUNK14_OFF + CHUNK14_SIZE]
    ids = [[rec_at(blob, x, y)[0] for x in range(60)] for y in range(60)]
    recs = [[rec_at(blob, x, y) for x in range(60)] for y in range(60)]

    special_cells = [(x, y) for y in range(60) for x in range(60) if ids[y][x] >= ID_SPECIAL]
    filled = []
    leftover_desc = []
    terrain_named = []
    blank_specials = []
    name_cells: dict[str, list[tuple[int, int]]] = defaultdict(list)
    name_ids: dict[str, set[int]] = defaultdict(set)

    for y in range(60):
        for x in range(60):
            lab = labels[y][x]
            tid = ids[y][x]
            if tid >= ID_SPECIAL:
                if is_placeholder(lab):
                    if lab.lower().startswith("desconhecido"):
                        leftover_desc.append((x, y, lab, tid))
                    blank_specials.append((x, y, tid, lab))
                else:
                    filled.append((x, y, lab, tid))
                    name_cells[lab].append((x, y))
                    name_ids[lab].add(tid)
            else:
                if lab and not is_placeholder(lab):
                    terrain_named.append((x, y, lab, tid))
                    name_cells[lab].append((x, y))
                    name_ids[lab].add(tid)

    print("=== fill counts ===")
    print(f"  special tiles (>=0x7D): {len(special_cells)}")
    print(f"  user-named specials:    {len(filled)}")
    print(f"  blank/Desconhecido spec:{len(blank_specials)}")
    print(f"  leftover Desconhecido*: {len(leftover_desc)}")
    print(f"  named terrain (<0x7D):  {len(terrain_named)}")
    print(f"  unique user names:      {len(name_cells)}")

    print("=== name -> ids + cells ===")
    for name in sorted(name_cells, key=lambda n: (-len(name_cells[n]), n.lower())):
        cells = name_cells[name]
        idset = sorted(name_ids[name])
        idhex = ",".join(f"0x{t:02X}" for t in idset)
        xs_ = [c[0] for c in cells]
        ys_ = [c[1] for c in cells]
        print(
            f"  {name!r:28} n={len(cells):4} ids=[{idhex}] "
            f"bbox=({min(xs_)},{min(ys_)})-({max(xs_)},{max(ys_)}) "
            f"w={max(xs_)-min(xs_)+1} h={max(ys_)-min(ys_)+1}"
        )
        if len(idset) > 1:
            by = Counter(ids[y][x] for x, y in cells)
            print(f"    MIXED IDS: { {hex(k): v for k, v in by.items()} }")

    # id consistency: same byte0 should have same name
    print("=== byte0 consistency (specials) ===")
    id_names: dict[int, Counter] = defaultdict(Counter)
    for y in range(60):
        for x in range(60):
            tid = ids[y][x]
            if tid < ID_SPECIAL:
                continue
            lab = labels[y][x]
            key = lab if not is_placeholder(lab) else "(blank/desc)"
            id_names[tid][key] += 1
    mixed_ids = []
    blank_ids = []
    named_ids = []
    for tid in sorted(id_names):
        c = id_names[tid]
        named_part = {k: v for k, v in c.items() if k != "(blank/desc)"}
        blank_n = c.get("(blank/desc)", 0)
        if named_part and blank_n:
            print(f"  PARTIAL 0x{tid:02X} named={named_part} blank={blank_n}")
        if len(named_part) > 1:
            mixed_ids.append(tid)
            print(f"  MIXED   0x{tid:02X} {dict(c)}")
        if named_part and not blank_n and len(named_part) == 1:
            named_ids.append((tid, next(iter(named_part)), sum(named_part.values())))
        if not named_part:
            blank_ids.append((tid, blank_n))

    print(f"  fully named ids: {len(named_ids)}")
    print(f"  fully blank ids: {len(blank_ids)}")
    print(f"  mixed-name ids:  {len(mixed_ids)}")
    print("  named table:")
    for tid, name, n in named_ids:
        print(f"    0x{tid:02X}  {name}  n={n}")
    print("  still-blank special ids:")
    for tid, n in blank_ids:
        cells = [(x, y) for y in range(60) for x in range(60) if ids[y][x] == tid]
        x0, y0 = cells[0]
        print(f"    0x{tid:02X}  n={n:4} first=({x0},{y0}) rec={recs[y0][x0].hex()}")

    # swapped-axis check: if user treated (x,y) as (y,x), names would smear
    print("=== axis / typo hints ===")
    # look at 2x2 same-id footprints named inconsistently
    for tid, name, n in named_ids:
        cells = [(x, y) for y in range(60) for x in range(60) if ids[y][x] == tid]
        xs_ = [c[0] for c in cells]
        ys_ = [c[1] for c in cells]
        w = max(xs_) - min(xs_) + 1
        h = max(ys_) - min(ys_) + 1
        if n in (4, 9, 16) and (w, h) not in ((2, 2), (3, 3), (4, 4)) and n == w * h:
            pass
        if n in (4, 9, 16) and {w, h} != {int(n**0.5)}:
            print(f"  odd footprint {name} 0x{tid:02X} n={n} bbox {w}x{h}")

    # typo-ish near names
    names = list(name_cells)
    for i, a in enumerate(names):
        for b in names[i + 1 :]:
            if a.lower() == b.lower() and a != b:
                print(f"  case-split: {a!r} vs {b!r}")
            if a.replace(" ", "") == b.replace(" ", "") and a != b:
                print(f"  space-split: {a!r} vs {b!r}")

    if leftover_desc:
        print("  leftover Desconhecido labels (user did not overwrite generator text):")
        by = Counter(lab for *_, lab, _ in leftover_desc)
        for lab, n in by.most_common():
            print(f"    {lab!r} n={n}")

    if terrain_named:
        print("  named terrain cells (unusual):")
        for x, y, lab, tid in terrain_named[:40]:
            print(f"    ({x},{y}) {lab!r} id=0x{tid:02X}")
        if len(terrain_named) > 40:
            print(f"    ... +{len(terrain_named)-40}")

    # actor26
    a26 = sav[CHUNK7_OFF : CHUNK7_OFF + CHUNK7_SIZE]
    print("=== actor26 chunk7 ===")
    live = 0
    types = Counter()
    for i in range(ACTOR_N):
        rec = a26[i * ACTOR_STRIDE : (i + 1) * ACTOR_STRIDE]
        occ = rec[0]
        if occ == 0:
            continue
        live += 1
        typ = rec[4]
        types[typ] += 1
        x, y = rec[6], rec[7]
        destx, desty = rec[0xE], rec[0xF]
        state = rec[0x12]
        spr = rec[1:4].hex()
        tile_off = int.from_bytes(rec[8:12], "little", signed=True)
        facing = rec[5]
        extra8a = rec[0x8A] if len(rec) > 0x8A else None
        print(
            f"  slot {i:2} occ={occ} type={typ} xy=({x},{y}) dest=({destx},{desty}) "
            f"state={state} face={facing} spr={spr} tile_off={tile_off} +8A={extra8a} "
            f"+1..+20={rec[0:21].hex()}"
        )
        if 0 <= x < 60 and 0 <= y < 60:
            tid = ids[y][x]
            print(f"       tile 0x{tid:02X} label={labels[y][x]!r} rec={recs[y][x].hex()} +7={recs[y][x][7]}")
    print(f"  live={live} types={dict(types)}")

    # walker pool — any with odd coords / type that might be ships?
    walk = sav[CHUNK8_OFF : CHUNK8_OFF + CHUNK8_SIZE]
    print("=== walker chunk8 (city 201x58) type histogram ===")
    wtypes = Counter()
    wlive = 0
    for i in range(WALKER_N):
        rec = walk[i * WALKER_STRIDE : (i + 1) * WALKER_STRIDE]
        if rec[0] == 0:
            continue
        wlive += 1
        wtypes[rec[4] if len(rec) > 4 else -1] += 1
    print(f"  live={wlive} types={dict(sorted(wtypes.items()))}")

    # tile +7 occupancy vs actors
    plus7 = [(x, y, recs[y][x][7]) for y in range(60) for x in range(60) if recs[y][x][7]]
    print(f"=== prov +7 slots n={len(plus7)} ===")
    for x, y, sl in plus7:
        print(f"  ({x},{y}) slot={sl} id=0x{ids[y][x]:02X} label={labels[y][x]!r}")

    # PL8 / strings in game dir
    print("=== game files matching boat/ship/sea/lane/prov/std ===")
    if GAME.is_dir():
        keys = ("boat", "ship", "sea", "lane", "prov", "std", "navy", "galley", "fleet")
        hits = []
        for p in GAME.rglob("*"):
            if not p.is_file():
                continue
            low = p.name.lower()
            if any(k in low for k in keys):
                hits.append(p)
        for p in sorted(hits, key=lambda z: z.name.lower()):
            print(f"  {p.name}  {p.stat().st_size}")
    else:
        print("  GAME dir missing")


if __name__ == "__main__":
    main()
