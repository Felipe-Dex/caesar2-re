#!/usr/bin/env python3
"""Export a 60×60 province-tile spreadsheet from SavChunk 14.

Does not copy the SAV. Does not invent city building names on province ids.
"""

from __future__ import annotations

import sys
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

SAV = Path(r"C:\Users\Felip\OneDrive\Games\Caesar2\Achea.sav\ACHEA23.SAV")
OUT_XLSX = ROOT / "findings" / "Achea_province.xlsx"

CHUNK14_OFF = 178395
CHUNK14_SIZE = 28800
MAP_W = MAP_H = 60
REC = 8
ID_SPECIAL = 0x7D  # FUN_00039032: byte0 < 0x7D terrain LUT, else building draw

# Only names pinned by EXE / Ghidra on the *province* map. Do not reuse city ids.
KNOWN: dict[int, str] = {
    0x92: "Invasão",  # actor26 type 2–5 on this id → walker type 3 (ghidra_walkers_tick.md)
}

FILLS = {
    "header": "1F4E79",
    "water": "6BAED6",
    "land": "E8E4D4",
    "land2": "D5C4A1",
    "feature": "A1D99B",
    "invasao": "E74C3C",
    "marker": "FDAE6B",
    "network": "B0B0B0",
    "built": "C5B0D5",
    "unknown": "FFF2A8",
}

THIN = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


def load_chunk14(path: Path) -> bytes:
    data = path.read_bytes()
    if len(data) != 225745:
        raise SystemExit(f"{path} size {len(data)} != 225745")
    blob = data[CHUNK14_OFF : CHUNK14_OFF + CHUNK14_SIZE]
    if len(blob) != CHUNK14_SIZE:
        raise SystemExit("chunk 14 short")
    return blob


def rec_at(blob: bytes, x: int, y: int) -> bytes:
    i = (y * MAP_W + x) * REC
    return blob[i : i + REC]


def class_of(tid: int) -> str:
    if tid in KNOWN:
        return "invasao"
    if tid < 0x18:
        return "water"
    if tid < ID_SPECIAL:
        return "land" if tid < 0x2A else "land2"
    if tid < 0x92:
        return "feature"
    if tid < 0xA0:
        return "marker"
    if tid < 0xB0:
        return "network"
    if tid < 0xD0:
        return "built"
    return "unknown"


def bbox_of(cells: list[tuple[int, int]]) -> dict:
    xs = [c[0] for c in cells]
    ys = [c[1] for c in cells]
    return {
        "xmin": min(xs),
        "xmax": max(xs),
        "ymin": min(ys),
        "ymax": max(ys),
        "w": max(xs) - min(xs) + 1,
        "h": max(ys) - min(ys) + 1,
        "n": len(cells),
    }


def main() -> None:
    blob = load_chunk14(SAV)
    nonzero = sum(1 for b in blob if b)
    ids = [[rec_at(blob, x, y)[0] for x in range(MAP_W)] for y in range(MAP_H)]
    recs = [[rec_at(blob, x, y) for x in range(MAP_W)] for y in range(MAP_H)]

    b0 = Counter(tid for row in ids for tid in row)
    special_ids = sorted(t for t in b0 if t >= ID_SPECIAL)
    by_id: dict[int, list[tuple[int, int]]] = {t: [] for t in special_ids}
    for y in range(MAP_H):
        for x in range(MAP_W):
            tid = ids[y][x]
            if tid >= ID_SPECIAL:
                by_id[tid].append((x, y))

    # Desconhecido N per unique id ≥ 0x7D (not per blob — same id = same type).
    # Order: first north-then-west cell of that id.
    numbered = sorted(
        special_ids,
        key=lambda t: (min(c[1] for c in by_id[t]), min(c[0] for c in by_id[t]), t),
    )
    n_of: dict[int, int] = {}
    name_of: dict[int, str] = {}
    n = 1
    for tid in numbered:
        if tid in KNOWN:
            name_of[tid] = KNOWN[tid]
            continue
        n_of[tid] = n
        name_of[tid] = f"Desconhecido {n}"
        n += 1
    unknown_ids = [t for t in numbered if t not in KNOWN]

    grid_text = [[""] * MAP_W for _ in range(MAP_H)]
    grid_class = [["land"] * MAP_W for _ in range(MAP_H)]
    for y in range(MAP_H):
        for x in range(MAP_W):
            tid = ids[y][x]
            grid_class[y][x] = class_of(tid)
            if tid >= ID_SPECIAL:
                grid_text[y][x] = name_of[tid]

    print(f"=== {SAV.name} chunk14 nonzero={nonzero}/{CHUNK14_SIZE} ===")
    print(f"  unique byte0={len(b0)}  specials(>=0x7D)={sum(b0[t] for t in special_ids)} tiles / {len(special_ids)} ids")
    print(f"  named={len(KNOWN)}  Desconhecido ids={len(unknown_ids)}")
    print("  byte0 top 15:")
    for tid, cnt in b0.most_common(15):
        print(f"    0x{tid:02X} n={cnt}")
    print("  corner / edge 1-tile specials:")
    for tid in special_ids:
        cells = by_id[tid]
        if len(cells) != 1:
            continue
        x, y = cells[0]
        edge = x <= 1 or y <= 1 or x >= 58 or y >= 58
        print(f"    0x{tid:02X} ({x},{y}) {name_of[tid]}  edge={edge}  rec={recs[y][x].hex()}")

    header_font = Font(name="Calibri", size=8, bold=True, color="FFFFFF")
    cell_font = Font(name="Calibri", size=7)
    header_fill = PatternFill("solid", fgColor=FILLS["header"])
    header_align = Alignment(horizontal="center", vertical="center")
    cell_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fill_cache = {k: PatternFill("solid", fgColor=v) for k, v in FILLS.items() if k != "header"}

    wb = Workbook()
    ws = wb.active
    ws.title = "mapa"

    ws.cell(1, 1, "y\\x")
    ws.cell(1, 1).font = header_font
    ws.cell(1, 1).fill = header_fill
    ws.cell(1, 1).alignment = header_align
    for x in range(MAP_W):
        cell = ws.cell(1, x + 2, x)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = THIN
    for y in range(MAP_H):
        cell = ws.cell(y + 2, 1, y)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = THIN
        for x in range(MAP_W):
            text = grid_text[y][x]
            cell = ws.cell(y + 2, x + 2, text if text else None)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = THIN
            cell.fill = fill_cache[grid_class[y][x]]

    ws.freeze_panes = "B2"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 14
    ws.column_dimensions["A"].width = 4.5
    for x in range(MAP_W):
        ws.column_dimensions[get_column_letter(x + 2)].width = 5.2
    for y in range(MAP_H):
        ws.row_dimensions[y + 2].height = 12

    lg = wb.create_sheet("legenda")
    lg["A1"] = "Como ler"
    lg["A1"].font = Font(name="Calibri", size=12, bold=True)
    notes = [
        "Save: ACHEA23.SAV. SavChunk 14 @ file 178395 = 60×60×8 @ VA 0xD94FC.",
        "(0,0) = ponta NORTE do losango da província (mesmo eixo do mapa da cidade).",
        "Números nas bordas são tile x=0..59 e y=0..59. Painéis congelados em B2.",
        "Célula sem texto = terreno (byte0 < 0x7D). Draw FUN_00039032: <0x7D LUT, ≥0x7D prédio.",
        "NÃO copiar ids da cidade (Tent/Casa/Plaza…). O limiar da província é 0x7D, não 0x78.",
        "Invasão = 0x92 (2×2): actor26 tipo 2–5 neste id spawna walker tipo 3 (bárbaro).",
        "Desconhecido N = um id ≥0x7D ainda sem nome. N é por id único, não por blob 4-conexo.",
        "Cores: azul=água-ish (<0x18), areia=terra, verde=feature REGIONS 0x7D–0x91, vermelho=invasão,",
        "laranja=marcador 0x93–0x9F, cinza=rede 0xA0–0xAF, roxo=outro built, amarelo=resto.",
    ]
    for i, line in enumerate(notes, start=2):
        lg.cell(i, 1, line)
        lg.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)

    hdr_row = 13
    headers = ["N", "hex", "count", "x0", "y0", "x1", "y1", "w", "h", "nome", "nota"]
    for c, h in enumerate(headers, start=1):
        cell = lg.cell(hdr_row, c, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    unk_yellow = PatternFill("solid", fgColor=FILLS["unknown"])
    named_red = PatternFill("solid", fgColor=FILLS["invasao"])

    legend_rows: list[tuple[int, int | str, dict]] = []
    for tid in numbered:
        bb = bbox_of(by_id[tid])
        legend_rows.append((tid, n_of.get(tid, ""), bb))
    legend_rows.sort(key=lambda z: (-z[2]["n"], z[2]["ymin"], z[2]["xmin"]))

    for i, (tid, nlab, bb) in enumerate(legend_rows):
        r = hdr_row + 1 + i
        edge = bb["n"] == 1 and (
            bb["xmin"] <= 1 or bb["ymin"] <= 1 or bb["xmax"] >= 58 or bb["ymax"] >= 58
        )
        nota = "Ghidra: spawn bárbaro"
        if tid not in KNOWN:
            nota = "canto/borda" if edge else "olha no jogo"
        vals = [
            nlab if nlab != "" else "—",
            f"0x{tid:02X}",
            bb["n"],
            bb["xmin"],
            bb["ymin"],
            bb["xmax"],
            bb["ymax"],
            bb["w"],
            bb["h"],
            name_of[tid],
            nota,
        ]
        fill = named_red if tid in KNOWN else unk_yellow
        for c, v in enumerate(vals, start=1):
            cell = lg.cell(r, c, v)
            cell.font = Font(name="Calibri", size=9)
            cell.fill = fill
    lg.auto_filter.ref = f"A{hdr_row}:K{hdr_row + len(legend_rows)}"
    lg.freeze_panes = f"A{hdr_row + 1}"
    for i, w in enumerate([6, 8, 8, 6, 6, 6, 6, 5, 5, 18, 22], start=1):
        lg.column_dimensions[get_column_letter(i)].width = w

    kn = wb.create_sheet("nomes")
    kn["A1"] = "id"
    kn["B1"] = "nome"
    kn["C1"] = "conf"
    for c in range(1, 4):
        kn.cell(1, c).font = header_font
        kn.cell(1, c).fill = header_fill
    kn.cell(2, 1, "0x92")
    kn.cell(2, 2, "Invasão (2×2) — actor26 pisa neste id → walker tipo 3")
    kn.cell(2, 3, "Ghidra HIGH")
    kn.cell(4, 1, "Menu província (C2MODEL custos, SEM id nesta planilha)")
    kn.merge_cells("A4:C4")
    kn.cell(5, 1, "—")
    kn.cell(5, 2, "Road / Wall / Fort / Work camp / Farm / Port / Warehouse / Shipyard / Trading post")
    kn.cell(5, 3, "custo only")
    kn.cell(6, 1, "C2.ENG [53]")
    kn.cell(6, 2, "Prov. Wall")
    kn.cell(6, 3, "label only")
    kn.column_dimensions["A"].width = 18
    kn.column_dimensions["B"].width = 78
    kn.column_dimensions["C"].width = 14
    kn.freeze_panes = "A2"

    hist = wb.create_sheet("byte0")
    hist["A1"] = "hex"
    hist["B1"] = "dec"
    hist["C1"] = "count"
    hist["D1"] = "classe"
    hist["E1"] = "nome"
    for c in range(1, 6):
        hist.cell(1, c).font = header_font
        hist.cell(1, c).fill = header_fill
    for i, tid in enumerate(sorted(b0), start=2):
        hist.cell(i, 1, f"0x{tid:02X}")
        hist.cell(i, 2, tid)
        hist.cell(i, 3, b0[tid])
        hist.cell(i, 4, class_of(tid) if tid >= ID_SPECIAL or tid < 0x18 else "terreno")
        hist.cell(i, 5, name_of.get(tid, ""))
    for i, w in enumerate([8, 6, 8, 12, 18], start=1):
        hist.column_dimensions[get_column_letter(i)].width = w
    hist.freeze_panes = "A2"
    hist.auto_filter.ref = f"A1:E{1 + len(b0)}"

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"=== WROTE {OUT_XLSX}  {OUT_XLSX.stat().st_size} B ===")
    print(f"unknown ids={len(unknown_ids)}  special tiles={sum(b0[t] for t in special_ids)}")


if __name__ == "__main__":
    main()
