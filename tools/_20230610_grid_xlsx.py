#!/usr/bin/env python3
"""80×80 named grid for 20230610.SAV. Achea v3 layout, Desconhecido N.

Does not copy the SAV.
"""

from __future__ import annotations

import sys
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / "tools"))

from app.city_map import (  # noqa: E402
    FLAG_PAD,
    FLAG_RIVER,
    ID_TERRAIN_MAX,
    MAP_H,
    MAP_W,
    load_city_from_sav,
)
from _achea_grid_xlsx import (  # noqa: E402
    CLASS_OF_NAME,
    FA_GOODS,
    FILLS,
    THIN,
    same_id_blobs,
)
from _20230610_parse import ACHEA_NAMED  # noqa: E402

SAV = Path(r"C:\Users\Felip\OneDrive\Games\Caesar2\20230610.SAV")
OUT_XLSX = ROOT / "findings" / "20230610_grid.xlsx"

# Extra class keys used by ACHEA_NAMED but missing from the Achea v3 map.
CLASS_OF_NAME.setdefault("Shrine 1", "worship")
CLASS_OF_NAME.setdefault("Shrine 2", "worship")
CLASS_OF_NAME.setdefault("Temple 1", "worship")
CLASS_OF_NAME.setdefault("Temple 2", "worship")
CLASS_OF_NAME.setdefault("Temple 3", "worship")
CLASS_OF_NAME.setdefault("Basilica 2", "worship")
CLASS_OF_NAME.setdefault("Aventine 1", "forum")
CLASS_OF_NAME.setdefault("Aventine 2", "forum")
CLASS_OF_NAME.setdefault("Aventine 3", "forum")
CLASS_OF_NAME.setdefault("Palatine 2", "forum")
CLASS_OF_NAME.setdefault("Palatine 4", "forum")
CLASS_OF_NAME.setdefault("Baths 3", "amenity")
CLASS_OF_NAME.setdefault("Market 1", "market")

HYP_NOTE = {
    0xCB: "Aqueduto ponta / stub (CITYFIXT type 0x11, +1=0x40, vizinho 0xCF). Nao e Well.",
}


def main() -> None:
    city = load_city_from_sav(SAV)
    ids = [[city.tile(x, y).terrain_id for x in range(MAP_W)] for y in range(MAP_H)]
    flags = [[city.tile(x, y).flags for x in range(MAP_W)] for y in range(MAP_H)]
    specials = [[city.tile(x, y).special for x in range(MAP_W)] for y in range(MAP_H)]
    blobs = same_id_blobs(ids)

    fa_name_at: dict[tuple[int, int], str] = {}
    for b in blobs:
        if b["id"] != 0xFA:
            continue
        nibs = {specials[cy][cx] & 0x0F for cx, cy in b["cells"] if specials[cy][cx] & 0x0F}
        goods = next(iter(nibs), 0) if len(nibs) == 1 else (max(nibs) if nibs else 0)
        name = FA_GOODS.get(goods, "Factory")
        for cx, cy in b["cells"]:
            fa_name_at[(cx, cy)] = name

    unnamed_blobs = [b for b in blobs if b["id"] not in ACHEA_NAMED]
    unnamed_blobs.sort(key=lambda b: (b["ymin"], b["xmin"], b["id"]))
    label_at: dict[tuple[int, int], str] = {}
    unk_meta: list[dict] = []
    for n, b in enumerate(unnamed_blobs, start=1):
        name = f"Desconhecido {n}"
        for x, y in b["cells"]:
            label_at[(x, y)] = name
        unk_meta.append({**b, "n_label": n, "name": name})

    grid_text = [[""] * MAP_W for _ in range(MAP_H)]
    grid_class = [["terrain"] * MAP_W for _ in range(MAP_H)]
    for y in range(MAP_H):
        for x in range(MAP_W):
            tid = ids[y][x]
            if tid == 0xFA:
                grid_text[y][x] = fa_name_at[(x, y)]
                grid_class[y][x] = CLASS_OF_NAME[grid_text[y][x]]
            elif tid in ACHEA_NAMED:
                grid_text[y][x] = ACHEA_NAMED[tid]
                grid_class[y][x] = CLASS_OF_NAME[ACHEA_NAMED[tid]]
            elif tid >= ID_TERRAIN_MAX:
                grid_text[y][x] = label_at[(x, y)]
                grid_class[y][x] = "unknown"
            elif flags[y][x] & FLAG_PAD:
                grid_text[y][x] = "Road"
                grid_class[y][x] = "road"
            elif flags[y][x] & FLAG_RIVER:
                grid_text[y][x] = "Rio"
                grid_class[y][x] = "water"

    header_font = Font(name="Calibri", size=8, bold=True, color="FFFFFF")
    cell_font = Font(name="Calibri", size=7)
    header_fill = PatternFill("solid", fgColor=FILLS["header"])
    header_align = Alignment(horizontal="center", vertical="center")
    cell_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fill_cache = {k: PatternFill("solid", fgColor=v) for k, v in FILLS.items() if k != "header"}

    wb = Workbook()
    ws = wb.active
    ws.title = "mapa"
    ws.cell(1, 1, "y\\x")
    ws.cell(1, 1).font = header_font
    ws.cell(1, 1).fill = header_fill
    ws.cell(1, 1).alignment = header_align
    for x in range(MAP_W):
        cell = ws.cell(1, x + 2, x)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = THIN
    for y in range(MAP_H):
        cell = ws.cell(y + 2, 1, y)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = THIN
        for x in range(MAP_W):
            text = grid_text[y][x]
            cell = ws.cell(y + 2, x + 2, text if text else None)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = THIN
            cell.fill = fill_cache[grid_class[y][x] if text else "terrain"]

    ws.freeze_panes = "B2"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 14
    ws.column_dimensions["A"].width = 4.5
    for x in range(MAP_W):
        ws.column_dimensions[get_column_letter(x + 2)].width = 5.2
    for y in range(MAP_H):
        ws.row_dimensions[y + 2].height = 12

    lg = wb.create_sheet("legenda")
    lg["A1"] = "Como ler"
    lg["A1"].font = Font(name="Calibri", size=12, bold=True)
    notes = [
        "20230610.SAV chunk 13. SAV fica no OneDrive — não copiado para o git.",
        "(0,0) = ponta NORTE. Iso: direita = x−y; baixo = x+y.",
        "Amarelo = id que a Achea ainda não tinha nome. N = blob 4-conexo, ordem norte-depois-oeste.",
        "Palatine 4 hyp = 0xB9 4×4 em (34,43)–(37,46), ao lado do único Basilica 0xAC (39,42)–(41,44).",
        "Não há 3×3 culto novo: Basilica 3 não apareceu como id distinto. Só 0xAC (Achea = 4º).",
        "Outros: 0xA7 Temple 2 · 0xE1 Baths 3 · 0xFC Market 1 · 0xCB aqueduto ponta (nao e Well).",
    ]
    for i, line in enumerate(notes, start=2):
        lg.cell(i, 1, line)
        lg.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)

    hdr_row = 10
    headers = ["N", "id_hex", "id_dec", "w", "h", "x0", "y0", "x1", "y1", "count", "nome", "nota"]
    for c, h in enumerate(headers, start=1):
        cell = lg.cell(hdr_row, c, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    unk_yellow = PatternFill("solid", fgColor=FILLS["unknown"])
    for i, b in enumerate(sorted(unk_meta, key=lambda z: (-z["n"], z["ymin"], z["xmin"]))):
        r = hdr_row + 1 + i
        vals = [
            b["n_label"],
            f"0x{b['id']:02X}",
            b["id"],
            b["w"],
            b["h"],
            b["xmin"],
            b["ymin"],
            b["xmax"],
            b["ymax"],
            b["n"],
            b["name"],
            HYP_NOTE.get(b["id"], "olha no jogo"),
        ]
        for c, v in enumerate(vals, start=1):
            cell = lg.cell(r, c, v)
            cell.font = Font(name="Calibri", size=9)
            cell.fill = unk_yellow
    lg.auto_filter.ref = f"A{hdr_row}:L{hdr_row + len(unk_meta)}"
    lg.freeze_panes = f"A{hdr_row + 1}"
    for i, w in enumerate([6, 10, 8, 5, 5, 6, 6, 6, 6, 8, 18, 56], start=1):
        lg.column_dimensions[get_column_letter(i)].width = w

    kn = wb.create_sheet("nomes")
    kn["A1"] = "id"
    kn["B1"] = "nome"
    kn["C1"] = "conf"
    for c in range(1, 4):
        kn.cell(1, c).font = header_font
        kn.cell(1, c).fill = header_fill
    known_rows = [
        ("0xA7", "Temple 2 (3× 2×2) — NOVO nesta save", "hyp HIGH"),
        ("0xB9", "Palatine 4 4×4 (34,43) — NOVO; 0xB7 ausente", "hyp HIGH"),
        ("0xE1", "Baths 3 (2× 2×2) — NOVO; +4 entre E0 e E2", "hyp HIGH"),
        ("0xFC", "Market 1 2×2 (39,39) — NOVO; +4 antes de FD", "hyp HIGH"),
        ("0xCB", "Aqueduto ponta / stub 1×1 (26,31). Nao e Well.", "20230610"),
        ("0xAC", "único Basilica: 3×3 (39,42). Achea = 4º. Sem id de 3º.", "Achea v3"),
        ("0xA6", "Temple 1", "Achea"),
        ("resto", "nomes Achea (Garden/Plaza/fóruns/etc.)", "Achea"),
    ]
    for i, row in enumerate(known_rows, start=2):
        for c, v in enumerate(row, start=1):
            kn.cell(i, c, v)
    kn.column_dimensions["A"].width = 18
    kn.column_dimensions["B"].width = 62
    kn.column_dimensions["C"].width = 14
    kn.freeze_panes = "A2"

    bld_ids = Counter(tid for row in ids for tid in row if tid >= ID_TERRAIN_MAX)
    named_id_set = {i for i in bld_ids if i in ACHEA_NAMED}
    unk_id_set = {i for i in bld_ids if i not in ACHEA_NAMED}
    named_tiles = sum(bld_ids[i] for i in named_id_set)
    unk_tiles = sum(bld_ids[i] for i in unk_id_set)
    name_tiles = Counter()
    for y in range(MAP_H):
        for x in range(MAP_W):
            if ids[y][x] >= ID_TERRAIN_MAX:
                name_tiles[grid_text[y][x]] += 1

    pg = wb.create_sheet("progress")
    pg["A1"] = "20230610_grid — ids fora do conjunto nomeado da Achea"
    pg["A1"].font = Font(name="Calibri", size=12, bold=True)
    summary = [
        ("tiles prédio (id≥0x78)", sum(bld_ids.values())),
        ("tiles nomeados (Achea)", named_tiles),
        ("tiles Desconhecido", unk_tiles),
        ("% tiles nomeados", f"{100.0 * named_tiles / sum(bld_ids.values()):.1f}%"),
        ("ids únicos prédio", len(bld_ids)),
        ("ids únicos nomeados", len(named_id_set)),
        ("ids únicos Desconhecido", len(unk_id_set)),
        ("blobs Desconhecido", len(unk_meta)),
    ]
    pg["A3"] = "métrica"
    pg["B3"] = "valor"
    for c in range(1, 3):
        pg.cell(3, c).font = header_font
        pg.cell(3, c).fill = header_fill
    for i, (lab, val) in enumerate(summary, start=4):
        pg.cell(i, 1, lab)
        pg.cell(i, 2, val)

    pg.cell(14, 1, "id Desconhecido")
    pg.cell(14, 2, "tiles")
    pg.cell(14, 3, "hex")
    pg.cell(14, 4, "hyp")
    for c in range(1, 5):
        pg.cell(14, c).font = header_font
        pg.cell(14, c).fill = header_fill
    for i, tid in enumerate(sorted(unk_id_set), start=15):
        pg.cell(i, 1, tid)
        pg.cell(i, 2, bld_ids[tid])
        pg.cell(i, 3, f"0x{tid:02X}")
        pg.cell(i, 4, HYP_NOTE.get(tid, ""))

    pg.cell(14, 6, "nome no mapa")
    pg.cell(14, 7, "tiles")
    pg.cell(14, 6).font = header_font
    pg.cell(14, 6).fill = header_fill
    pg.cell(14, 7).font = header_font
    pg.cell(14, 7).fill = header_fill
    for i, (nm, n) in enumerate(name_tiles.most_common(), start=15):
        pg.cell(i, 6, nm)
        pg.cell(i, 7, n)
    pg.column_dimensions["A"].width = 32
    pg.column_dimensions["B"].width = 14
    pg.column_dimensions["C"].width = 10
    pg.column_dimensions["D"].width = 64
    pg.column_dimensions["F"].width = 22
    pg.column_dimensions["G"].width = 10
    pg.freeze_panes = "A15"

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"WROTE {OUT_XLSX}  {OUT_XLSX.stat().st_size} B")
    print(f"unknown ids={sorted(f'0x{i:02X}' for i in unk_id_set)}  blobs={len(unk_meta)}")
    for b in unk_meta:
        print(
            f"  N={b['n_label']} 0x{b['id']:02X} {b['w']}x{b['h']} "
            f"({b['xmin']},{b['ymin']})-({b['xmax']},{b['ymax']})  {HYP_NOTE.get(b['id'], '')}"
        )


if __name__ == "__main__":
    main()
