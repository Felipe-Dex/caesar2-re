#!/usr/bin/env python3
"""Mark live province actor26 from ACHEA23 onto a new Excel grid.

Same coord convention as findings/Achea_province.xlsx (row 1 = x, col A = y).
Reads user labels from that workbook. Does not write it. Does not copy the SAV.
"""

from __future__ import annotations

import struct
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / "tools"))

from _province_xlsx import FILLS, ID_SPECIAL, THIN, class_of  # noqa: E402

SAV = Path(r"C:\Users\Felip\OneDrive\Games\Caesar2\Achea.sav\ACHEA23.SAV")
LABELS_XLSX = ROOT / "findings" / "Achea_province.xlsx"
OUT_XLSX = ROOT / "findings" / "Achea_province_walkers.xlsx"

CHUNK7_OFF = 16
CHUNK7_SIZE = 4550
CHUNK14_OFF = 178395
CHUNK14_SIZE = 28800
MAP_W = MAP_H = 60
TILE_REC = 8
ACTOR_N = 26
ACTOR_STRIDE = 175

# C2.ENG extra-pool after official [44] "Formed " — NOT city [66]+type.
# Type 1 sits on the Fort; type 6 is the ship handler. 2–5 / 7–8 absent here.
TYPE_LABEL = {
    1: "Prima Cohors",
    2: "Enemy Army?",
    3: "Barbarians?",
    4: "Enemy Army?",
    5: "Enemy Army?",
    6: "Merchant Ship",
    7: "Enemy Ship?",
    8: "Barbarian Ship?",
}

TYPE_HINT = {
    1: "Painel Prima Cohors / Formed year. Clique o SPRITE no Fort. Ver findings/province_actors.md.",
    2: "exército terrestre — pisa Your City 0x92 → walker tipo 3. Zero nesta save.",
    3: "exército terrestre (0x92). Zero nesta save.",
    4: "exército terrestre (0x92). Zero nesta save.",
    5: "exército terrestre (0x92). Zero nesta save.",
    6: "ACHEA23: Merchant Ship. Silk = slot 3 (44,18). EXE 6=Merchant / 7=Enemy / 8=Barbarian.",
    7: "EXE: Enemy Ship. handler RET — sem AI neste build. Zero nesta save.",
    8: "EXE: Barbarian Ship. handler RET — sem AI neste build. Zero nesta save.",
}

TYPE_FILL = {
    1: "E76F51",
    2: "264653",
    3: "2A9D8F",
    4: "E9C46A",
    5: "F4A261",
    6: "4C8BF5",
    7: "C0392B",
    8: "7F8C8D",
}

SKIP_NEAR = {
    "",
    "terreno",
    "terrain",
}

HEADER_FONT = Font(name="Calibri", size=8, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor=FILLS["header"])
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
CELL_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
QUERY_BORDER = Border(
    left=Side(style="medium", color="1A1A1A"),
    right=Side(style="medium", color="1A1A1A"),
    top=Side(style="medium", color="1A1A1A"),
    bottom=Side(style="medium", color="1A1A1A"),
)
THICK_YELLOW = PatternFill("solid", fgColor="FFE066")
FACING = ("N", "NE", "E", "SE", "S", "SW", "W", "NW")


def i8(b: int) -> int:
    return b - 256 if b >= 128 else b


def rec_at(blob: bytes, x: int, y: int) -> bytes:
    return blob[(y * MAP_W + x) * TILE_REC : (y * MAP_W + x) * TILE_REC + TILE_REC]


def excel_cell(x: int, y: int) -> str:
    return f"{get_column_letter(x + 2)}{y + 2}"


def is_placeholder(lab: str) -> bool:
    if not lab:
        return True
    low = lab.lower().strip()
    if low in SKIP_NEAR:
        return True
    return low.startswith("desconhecido")


def load_user_labels() -> list[list[str]]:
    wb = load_workbook(LABELS_XLSX, data_only=True)
    ws = wb["mapa"]
    return [
        [str(ws.cell(y + 2, x + 2).value or "").strip() for x in range(MAP_W)]
        for y in range(MAP_H)
    ]


def tile_caption(tid: int, lab: str) -> str:
    if not is_placeholder(lab):
        return lab
    if tid < ID_SPECIAL:
        return ""
    return f"id0x{tid:02X}"


def nearby_named(
    blob: bytes, labels: list[list[str]], x: int, y: int, radius: int = 12
) -> list[str]:
    hits: list[tuple[int, str]] = []
    seen: set[str] = set()
    for dy in range(-radius, radius + 1):
        for dx in range(-radius, radius + 1):
            nx, ny = x + dx, y + dy
            if not (0 <= nx < MAP_W and 0 <= ny < MAP_H):
                continue
            lab = labels[ny][nx]
            if is_placeholder(lab):
                continue
            key = lab.split("(")[0].strip()
            if key in seen:
                continue
            seen.add(key)
            hits.append((abs(dx) + abs(dy), f"{lab} ({nx},{ny})"))
    hits.sort()
    return [h[1] for h in hits[:5]]


def dest_note(dx: int, dy: int) -> str:
    if not (0 <= dx < MAP_W and 0 <= dy < MAP_H):
        return "nenhum (fora 0–59)"
    return f"({dx},{dy})"


def parse_actors(sav: bytes, blob: bytes, labels: list[list[str]]) -> list[dict]:
    pool = sav[CHUNK7_OFF : CHUNK7_OFF + CHUNK7_SIZE]
    rows: list[dict] = []
    for slot in range(ACTOR_N):
        rec = pool[slot * ACTOR_STRIDE : (slot + 1) * ACTOR_STRIDE]
        occ = rec[0]
        if occ == 0:
            continue
        typ = rec[4]
        x, y = i8(rec[6]), i8(rec[7])
        xf, yf = rec[0xC], rec[0xD]
        dest_x, dest_y = i8(rec[0xE]), i8(rec[0xF])
        state = rec[0x12]
        facing = rec[5]
        tile_off = struct.unpack_from("<i", rec, 8)[0]
        expect = (y * MAP_W + x) * TILE_REC if 0 <= x < MAP_W and 0 <= y < MAP_H else -1
        tid = rec_at(blob, x, y)[0] if expect >= 0 else -1
        plus7 = rec_at(blob, x, y)[7] if expect >= 0 else -1
        lab = labels[y][x] if expect >= 0 else ""
        rows.append(
            {
                "slot": slot,
                "occ": occ,
                "type": typ,
                "x": x,
                "y": y,
                "x_frac": xf,
                "y_frac": yf,
                "dest_x": dest_x,
                "dest_y": dest_y,
                "dest": dest_note(dest_x, dest_y),
                "state": state,
                "facing": facing,
                "face": FACING[facing] if 0 <= facing < 8 else str(facing),
                "sprite": rec[1],
                "tile_off": tile_off,
                "tile_off_ok": tile_off == expect,
                "tid": tid,
                "plus7": plus7,
                "on": tile_caption(tid, lab) or f"terreno 0x{tid:02X}",
                "near": nearby_named(blob, labels, x, y),
                "excel": excel_cell(x, y),
            }
        )
    rows.sort(key=lambda r: (r["type"], r["y"], r["x"], r["slot"]))
    return rows


def pick_query(rows: list[dict]) -> list[dict]:
    """One easy cell per live type. Parked ship over the walking one."""
    chosen: list[dict] = []
    by_type: dict[int, list[dict]] = defaultdict(list)
    for row in rows:
        by_type[row["type"]].append(row)
    for typ in sorted(by_type):
        candidates = by_type[typ]
        if typ == 6:
            candidates = sorted(
                candidates,
                key=lambda r: (0 if r["dest_x"] == r["x"] and r["dest_y"] == r["y"] else 1, r["y"], r["x"]),
            )
        copy = dict(candidates[0])
        copy["query_n"] = len(chosen) + 1
        chosen.append(copy)
        if typ == 6 and len(candidates) > 1:
            alt = dict(candidates[1])
            alt["query_n"] = 0
            alt["query_alt"] = True
            chosen.append(alt)
    return chosen


def write_map(
    wb: Workbook, blob: bytes, labels: list[list[str]], rows: list[dict], query: list[dict]
) -> None:
    ws = wb.active
    ws.title = "mapa"
    by_xy: dict[tuple[int, int], list[dict]] = defaultdict(list)
    for row in rows:
        by_xy[(row["x"], row["y"])].append(row)
    query_xy = {(q["x"], q["y"]): q for q in query if q.get("query_n")}
    pale = {k: PatternFill("solid", fgColor=v) for k, v in FILLS.items() if k != "header"}
    type_fills = {t: PatternFill("solid", fgColor=c) for t, c in TYPE_FILL.items()}
    cell_font = Font(name="Calibri", size=7)
    walker_font = Font(name="Calibri", size=8, bold=True)
    query_font = Font(name="Calibri", size=8, bold=True, color="1A1A1A")

    ws.cell(1, 1, "y\\x")
    ws.cell(1, 1).font = HEADER_FONT
    ws.cell(1, 1).fill = HEADER_FILL
    ws.cell(1, 1).alignment = HEADER_ALIGN
    for x in range(MAP_W):
        cell = ws.cell(1, x + 2, x)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN
    for y in range(MAP_H):
        cell = ws.cell(y + 2, 1, y)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN
        for x in range(MAP_W):
            here = by_xy.get((x, y), [])
            q = query_xy.get((x, y))
            tid = rec_at(blob, x, y)[0]
            lab = labels[y][x]
            cap = tile_caption(tid, lab)
            cell = ws.cell(y + 2, x + 2)
            cell.alignment = CELL_ALIGN
            cell.border = THIN
            if here:
                tags = "+".join(f"t{r['type']}" for r in here)
                if q:
                    cell.value = f"★t{q['type']}"
                    cell.font = query_font
                    cell.fill = THICK_YELLOW
                    cell.border = QUERY_BORDER
                else:
                    cell.value = tags
                    cell.font = walker_font
                    cell.fill = type_fills.get(here[0]["type"], pale["built"])
                slots = ", ".join(f"A{r['slot']} t{r['type']}" for r in here)
                extra = f"Query #{q['query_n']} — este primeiro" if q else ""
                cell.comment = Comment(
                    f"{slots}\n({x},{y}) em {here[0]['on']}\n"
                    f"perto: {', '.join(here[0]['near'][:3]) or '—'}\n{extra}".strip(),
                    "achea_province_walkers",
                )
                cell.comment.width = 240
                cell.comment.height = 90
            else:
                cell.value = cap or None
                cell.font = cell_font
                cell.fill = pale.get(class_of(tid), pale["land"])

    ws.freeze_panes = "B2"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 14
    ws.column_dimensions["A"].width = 4.5
    for x in range(MAP_W):
        ws.column_dimensions[get_column_letter(x + 2)].width = 5.4
    for y in range(MAP_H):
        ws.row_dimensions[y + 2].height = 12


def write_legend(wb: Workbook, rows: list[dict], query: list[dict]) -> None:
    lg = wb.create_sheet("legenda")
    lg["A1"] = "Actors da província — ACHEA23.SAV chunk 7 (26 × 175)"
    lg["A1"].font = Font(name="Calibri", size=12, bold=True)
    notes = [
        "NÃO são os 91 walkers da cidade (chunk 8). Pool separado: actor26, SavChunk 7, desenho MY_STDS.PL8.",
        "Mesma convenção do Achea_province.xlsx: linha 1 = x, coluna A = y. Célula Excel = tile (x, y).",
        "(0,0) = ponta NORTE do losango. x/y em +6/+7 são tile inteiro 0–59; +0xC/+0xD = frac (sub-tile do sprite).",
        "Tipo cidade ≠ tipo província. Cidade t1 = Forum Clerk. Província t1 = Cohort no Fort.",
        "C2.ENG [66] é título de walker da cidade. Navio/exército estão no pool extra depois de [44] Formed.",
    ]
    for i, line in enumerate(notes, start=2):
        lg.cell(i, 1, line)
        lg.merge_cells(start_row=i, start_column=1, end_row=i, end_column=14)

    query_key = {(q["slot"], q["type"]) for q in query if q.get("query_n")}
    hdr_row = 8
    headers = [
        "slot",
        "tipo",
        "nome (C2.ENG)",
        "x",
        "y",
        "Excel",
        "estado",
        "facing",
        "dest",
        "frac_x",
        "frac_y",
        "no tile",
        "perto",
        "tile+7",
        "Query?",
    ]
    for c, h in enumerate(headers, start=1):
        cell = lg.cell(hdr_row, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    type_fills = {t: PatternFill("solid", fgColor=c) for t, c in TYPE_FILL.items()}
    body = Font(name="Calibri", size=9)
    for i, row in enumerate(rows):
        r = hdr_row + 1 + i
        is_q = (row["slot"], row["type"]) in query_key
        vals = [
            row["slot"],
            row["type"],
            TYPE_LABEL.get(row["type"], f"t{row['type']}"),
            row["x"],
            row["y"],
            row["excel"],
            row["state"],
            row["face"],
            row["dest"],
            row["x_frac"],
            row["y_frac"],
            row["on"],
            ", ".join(row["near"][:3]),
            row["plus7"],
            "★ PRIMEIRO" if is_q else "",
        ]
        fill = THICK_YELLOW if is_q else type_fills.get(row["type"], PatternFill("solid", fgColor="DDDDDD"))
        for c, v in enumerate(vals, start=1):
            cell = lg.cell(r, c, v if v != "" else None)
            cell.font = body
            cell.fill = fill
    lg.auto_filter.ref = f"A{hdr_row}:O{hdr_row + len(rows)}"
    lg.freeze_panes = f"A{hdr_row + 1}"
    widths = [7, 7, 16, 5, 5, 8, 8, 8, 22, 8, 8, 28, 48, 8, 12]
    for i, w in enumerate(widths, start=1):
        lg.column_dimensions[get_column_letter(i)].width = w


def write_query(wb: Workbook, query: list[dict], rows: list[dict]) -> None:
    qs = wb.create_sheet("query", 0)
    qs["A1"] = "Query estes na PROVÍNCIA (não na cidade)"
    qs["A1"].font = Font(name="Calibri", size=14, bold=True)
    howto = [
        "1. Abra ACHEA23.SAV, vá ao mapa da província, pause.",
        "2. Clique o SPRITE (coorte / navio), não o prédio de baixo. Query da cidade não serve.",
        "3. Título da cidade usa C2.ENG [66]+tipo. Aqui o título é outra tabela (Cohort / Merchant Ship / …).",
        "Navio em movimento (8,43) anda para a Sea Lane Campania — se sumiu, use o parked.",
        "Tipos 2–5 (exército inimigo) e 7–8 não existem nesta save.",
    ]
    for i, line in enumerate(howto, start=2):
        qs.cell(i, 1, line)
        qs.merge_cells(start_row=i, start_column=1, end_row=i, end_column=10)

    counts: dict[int, int] = defaultdict(int)
    for row in rows:
        counts[row["type"]] += 1

    qs.cell(8, 1, "resumo desta save")
    qs.cell(8, 1).font = Font(name="Calibri", size=11, bold=True)
    qs.cell(9, 1, f"{len(rows)} actors vivos (pool 26). Cidade tem ~91 pessoas no chunk 8 — outro pool.")
    qs.merge_cells("A9:F9")
    r = 10
    for c, h in enumerate(("tipo", "qtd", "nome", "exemplo (x,y)", "Excel", "hint"), start=1):
        qs.cell(r, c, h)
        qs.cell(r, c).font = HEADER_FONT
        qs.cell(r, c).fill = HEADER_FILL
    type_fills = {t: PatternFill("solid", fgColor=c) for t, c in TYPE_FILL.items()}
    body = Font(name="Calibri", size=10)
    first_of: dict[int, dict] = {}
    for q in query:
        first_of.setdefault(q["type"], q)
    for i, typ in enumerate((1, 2, 3, 4, 5, 6, 7, 8)):
        rr = 11 + i
        ex = first_of.get(typ)
        e1 = f"({ex['x']},{ex['y']})" if ex else "—"
        xl = ex["excel"] if ex else "—"
        vals = [typ, counts.get(typ, 0), TYPE_LABEL[typ], e1, xl, TYPE_HINT[typ]]
        fill = type_fills.get(typ, PatternFill("solid", fgColor="DDDDDD"))
        for c, v in enumerate(vals, start=1):
            cell = qs.cell(rr, c, v)
            cell.font = body
            cell.fill = fill

    hdr = 21
    qs.cell(hdr, 1, "células para clicar")
    qs.cell(hdr, 1).font = Font(name="Calibri", size=11, bold=True)
    qs.merge_cells(start_row=hdr, start_column=1, end_row=hdr, end_column=10)
    headers = ["#", "tipo", "x", "y", "Excel", "slot", "estado", "no tile", "perto", "como achar"]
    for c, h in enumerate(headers, start=1):
        cell = qs.cell(hdr + 1, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    stars = [q for q in query if q.get("query_n")]
    alts = [q for q in query if q.get("query_alt")]
    for i, q in enumerate(stars + alts):
        is_first = bool(q.get("query_n"))
        tag = f"★{q['query_n']}" if is_first else "alt"
        near = ", ".join(q["near"][:2]) or q["on"]
        how = f"{q['on']} em ({q['x']},{q['y']}) — {near}"
        if not is_first:
            how = f"reserva t{q['type']} se o ★ já andou. {how}"
        rr = hdr + 2 + i
        vals = [
            tag,
            q["type"],
            q["x"],
            q["y"],
            q["excel"],
            q["slot"],
            q["state"],
            q["on"],
            ", ".join(q["near"][:2]),
            how,
        ]
        fill = THICK_YELLOW if is_first else type_fills[q["type"]]
        for c, v in enumerate(vals, start=1):
            cell = qs.cell(rr, c, v)
            cell.font = Font(name="Calibri", size=10, bold=is_first)
            cell.fill = fill
            cell.border = QUERY_BORDER if is_first else THIN

    widths = [8, 7, 5, 5, 8, 7, 8, 28, 42, 62]
    for i, w in enumerate(widths, start=1):
        qs.column_dimensions[get_column_letter(i)].width = w
    qs.freeze_panes = "A23"


def main() -> None:
    if not SAV.is_file():
        raise SystemExit(f"missing SAV: {SAV}")
    if not LABELS_XLSX.is_file():
        raise SystemExit(f"missing labels: {LABELS_XLSX}")
    sav = SAV.read_bytes()
    if len(sav) != 225745:
        raise SystemExit(f"SAV size {len(sav)} != 225745")
    blob = sav[CHUNK14_OFF : CHUNK14_OFF + CHUNK14_SIZE]
    labels = load_user_labels()
    rows = parse_actors(sav, blob, labels)
    query = pick_query(rows)

    wb = Workbook()
    write_map(wb, blob, labels, rows, query)
    write_legend(wb, rows, query)
    write_query(wb, query, rows)
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"=== WROTE {OUT_XLSX}  {OUT_XLSX.stat().st_size} B ===")
    print(f"live actors: {len(rows)}")
    counts: dict[int, int] = defaultdict(int)
    for row in rows:
        counts[row["type"]] += 1
    print("types:", dict(sorted(counts.items())))
    print("Query first:")
    for q in query:
        mark = "*" if q.get("query_n") else " "
        print(
            f"  {mark} t{q['type']} ({q['x']},{q['y']}) Excel {q['excel']} "
            f"A{q['slot']} on {q['on']} dest {q['dest']}"
        )


if __name__ == "__main__":
    main()
