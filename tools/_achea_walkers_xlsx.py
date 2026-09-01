#!/usr/bin/env python3
"""Mark live city walkers from ACHEA23 onto an Excel grid.

Same coord convention as findings/Achea_grid_v3.xlsx (row 1 = x, col A = y).
Does not copy the SAV. Does not overwrite Achea_grid_v3 building sheets.
"""

from __future__ import annotations

import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / "tools"))

from app.city_map import (  # noqa: E402
    FLAG_PAD,
    FLAG_RIVER,
    ID_TERRAIN_MAX,
    MAP_H,
    MAP_W,
    load_city_from_sav,
)
from app.walkers import live_walkers, load_walkers_from_sav  # noqa: E402
from _achea_grid_xlsx import (  # noqa: E402
    CLASS_OF_NAME,
    FA_GOODS,
    FILLS,
    KNOWN,
    THIN,
)

SAV = Path(r"C:\Users\Felip\OneDrive\Games\Caesar2\Achea.sav\ACHEA23.SAV")
OUT_XLSX = ROOT / "findings" / "Achea_walkers.xlsx"

# Query titles from Achea Q&A (findings/achea_walkers.md). 3/7 still spawn-only.
TYPE_LABEL = {
    1: "Forum Clerk",
    2: "Market Trader",
    3: "bárbaro",
    4: "Soldier",
    5: "Vigile",
    6: "Worker",
    7: "revoltoso",
}

# Person + cropped quote for the slots the user Query'd (left of the ★ cells).
QUERY_HITS = {
    34: ("Maelius Piscator", "We have excellen…"),
    67: ("Ennius Lentulus", "We have enough e…"),
    20: ("Caelius Clodius", "We have good pat…"),
    15: ("Aemilius Calvus", "This is a very law…"),
    90: ("Gaius Pernix", "We could do with…"),
    68: ("Iunius Maior", "We could do with…"),
}

TYPE_HINT = {
    1: "spawna de indústria / fórum (0xAE–0xB9); state 3 pinta cobertura 0x0C",
    2: "spawna de Market 0xFC–0xFF; state 4 pinta cobertura 0xC0; home = market",
    3: "bárbaro — nenhum nesta save",
    4: "spawna de Barracks 0xE4 / Tower 0xBF; state 7",
    5: "spawna de Praefecture 0xE3; state 8",
    6: "segundo emissor do mesmo bloco que type 2; home = Factory/Winery/Lead Works",
    7: "revoltoso — nenhum nesta save",
}

TYPE_FILL = {
    1: "F4A261",
    2: "2A9D8F",
    3: "264653",
    4: "E76F51",
    5: "E9C46A",
    6: "9B5DE5",
    7: "C0392B",
}

# Civic column the user already named (Achea_grid_v3).
LANDMARKS = (
    (71, 13, "Colosseum"),
    (71, 17, "Basilica"),
    (71, 10, "Basilica"),
    (71, 20, "Palatine"),
    (71, 25, "C.Maximus"),
    (70, 1, "Gate"),
    (73, 3, "Temple"),
    (62, 2, "Circus"),
    (74, 6, "Barracks"),
)

FACING = ("N", "NE", "E", "SE", "S", "SW", "W", "NW")

SKIP_NEAR = {"", "Road", "Rio", "terreno", "Tent", "Casa"}

HEADER_FONT = Font(name="Calibri", size=8, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor=FILLS["header"])
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
CELL_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
QUERY_BORDER = Border(
    left=Side(style="medium", color="1A1A1A"),
    right=Side(style="medium", color="1A1A1A"),
    top=Side(style="medium", color="1A1A1A"),
    bottom=Side(style="medium", color="1A1A1A"),
)
THICK_YELLOW = PatternFill("solid", fgColor="FFE066")


def tile_name(city, x: int, y: int) -> str:
    t = city.tile(x, y)
    tid = t.terrain_id
    if tid == 0xFA:
        return FA_GOODS.get(t.special & 0x0F, "Factory")
    if tid in KNOWN:
        return KNOWN[tid]
    if tid >= ID_TERRAIN_MAX:
        return f"id0x{tid:02X}"
    if t.flags & FLAG_PAD:
        return "Road"
    if t.flags & FLAG_RIVER:
        return "Rio"
    return "terreno"


def tile_class(city, x: int, y: int) -> str:
    name = tile_name(city, x, y)
    if name in CLASS_OF_NAME:
        return CLASS_OF_NAME[name]
    if name.startswith("id0x"):
        return "unknown"
    if name == "terreno":
        return "terrain"
    return "terrain"


def nearby_named(city, x: int, y: int, radius: int = 2) -> list[str]:
    seen: list[str] = []
    for dy in range(-radius, radius + 1):
        for dx in range(-radius, radius + 1):
            nx, ny = x + dx, y + dy
            if not (0 <= nx < MAP_W and 0 <= ny < MAP_H):
                continue
            name = tile_name(city, nx, ny)
            if name in SKIP_NEAR or name.startswith("Casa"):
                continue
            if name not in seen:
                seen.append(name)
    return seen


def home_xy(off: int) -> tuple[int, int] | None:
    if off <= 0 or off % 20:
        return None
    idx = off // 20
    x, y = idx % MAP_W, idx // MAP_W
    if 0 <= x < MAP_W and 0 <= y < MAP_H:
        return x, y
    return None


def excel_cell(x: int, y: int) -> str:
    return f"{get_column_letter(x + 2)}{y + 2}"


def is_core(x: int, y: int) -> bool:
    return x >= 35 and y <= 50


def pick_query_examples(rows: list[dict], per_type: int = 2) -> list[dict]:
    """Isolated walkers near the civic column the user already knows."""

    def score(row: dict) -> tuple:
        x, y = row["x"], row["y"]
        land = min(((x - lx) ** 2 + (y - ly) ** 2, name) for lx, ly, name in LANDMARKS)
        on_open = 0 if row["on"] in ("Road", "Plaza 1", "Plaza", "Plaza est") else 1
        core = 0 if is_core(x, y) else 1
        # Civic column x≈70 is the stack they already Query'd.
        col = abs(x - 70)
        return (row["stack"] > 1, core, on_open, col, land[0], y, x, row["slot"])

    chosen: list[dict] = []
    by_type: dict[int, list[dict]] = defaultdict(list)
    for row in rows:
        by_type[row["type"]].append(row)
    for typ in sorted(by_type):
        ranked = sorted(by_type[typ], key=score)
        for i, row in enumerate(ranked[:per_type]):
            copy = dict(row)
            copy["query_rank"] = i + 1
            copy["query_n"] = len(chosen) + 1
            chosen.append(copy)
    return chosen


def build_rows(city, walkers) -> list[dict]:
    by_xy: dict[tuple[int, int], list] = defaultdict(list)
    for w in walkers:
        by_xy[(w.x, w.y)].append(w)

    rows: list[dict] = []
    for w in walkers:
        hx = home_xy(w.home_off)
        near = nearby_named(city, w.x, w.y)
        rows.append(
            {
                "slot": w.slot,
                "type": w.type,
                "x": w.x,
                "y": w.y,
                "state": w.state,
                "facing": w.facing,
                "face": FACING[w.facing] if 0 <= w.facing < 8 else str(w.facing),
                "dest_x": w.dest_x,
                "dest_y": w.dest_y,
                "life": w.life_phase,
                "on": tile_name(city, w.x, w.y),
                "near": near,
                "home": hx,
                "home_name": tile_name(city, *hx) if hx else "",
                "stack": len(by_xy[(w.x, w.y)]),
                "excel": excel_cell(w.x, w.y),
            }
        )
    rows.sort(key=lambda r: (r["type"], r["y"], r["x"], r["slot"]))
    return rows


def write_map(wb: Workbook, city, rows: list[dict], query: list[dict]) -> None:
    if "mapa" in wb.sheetnames:
        ws = wb["mapa"]
    else:
        ws = wb.active
        ws.title = "mapa"

    by_xy: dict[tuple[int, int], list[dict]] = defaultdict(list)
    for row in rows:
        by_xy[(row["x"], row["y"])].append(row)
    query_xy = {(q["x"], q["y"]): q for q in query}

    pale = {k: PatternFill("solid", fgColor=v) for k, v in FILLS.items() if k != "header"}
    type_fills = {t: PatternFill("solid", fgColor=c) for t, c in TYPE_FILL.items()}
    cell_font = Font(name="Calibri", size=7)
    walker_font = Font(name="Calibri", size=8, bold=True)
    query_font = Font(name="Calibri", size=8, bold=True, color="1A1A1A")

    ws.cell(1, 1, "y\\x")
    ws.cell(1, 1).font = HEADER_FONT
    ws.cell(1, 1).fill = HEADER_FILL
    ws.cell(1, 1).alignment = HEADER_ALIGN
    for x in range(MAP_W):
        cell = ws.cell(1, x + 2, x)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN
    for y in range(MAP_H):
        cell = ws.cell(y + 2, 1, y)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN
        for x in range(MAP_W):
            here = by_xy.get((x, y), [])
            q = query_xy.get((x, y))
            name = tile_name(city, x, y)
            cls = tile_class(city, x, y)
            cell = ws.cell(y + 2, x + 2)
            cell.alignment = CELL_ALIGN
            cell.border = THIN
            if here:
                tags = "+".join(f"t{r['type']}" for r in here)
                if q:
                    cell.value = f"★t{q['type']}"
                    cell.font = query_font
                    cell.fill = THICK_YELLOW
                    cell.border = QUERY_BORDER
                else:
                    cell.value = tags
                    cell.font = walker_font
                    if len({r["type"] for r in here}) == 1:
                        cell.fill = type_fills[here[0]["type"]]
                    else:
                        cell.fill = PatternFill("solid", fgColor="C39BD3")
                slots = ", ".join(f"W{r['slot']} t{r['type']}" for r in here)
                extra = f"Query #{q['query_n']} — este primeiro" if q else ""
                cell.comment = Comment(
                    f"{slots}\n({x},{y}) em {name}\n"
                    f"perto: {', '.join(here[0]['near'][:4]) or '—'}\n{extra}".strip(),
                    "achea_walkers",
                )
                cell.comment.width = 220
                cell.comment.height = 80
            else:
                cell.value = name if name not in ("terreno", "") else None
                cell.font = cell_font
                cell.fill = pale.get(cls, pale["terrain"])

    ws.freeze_panes = "B2"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 14
    ws.column_dimensions["A"].width = 4.5
    for x in range(MAP_W):
        ws.column_dimensions[get_column_letter(x + 2)].width = 5.4
    for y in range(MAP_H):
        ws.row_dimensions[y + 2].height = 12


def write_legend(wb: Workbook, rows: list[dict], query: list[dict]) -> None:
    lg = wb.create_sheet("legenda")
    lg["A1"] = "Walkers vivos — ACHEA23.SAV (fonte da verdade)"
    lg["A1"].font = Font(name="Calibri", size=12, bold=True)
    notes = [
        "Mesma convenção do Achea_grid_v3: linha 1 = x, coluna A = y. Célula Excel = (x, y) do jogo.",
        "(0,0) = ponta NORTE do losango. Não é o canto superior-esquerdo da tela.",
        "Query no DosBox usa o tile (x, y). Se você costuma escrever (y,x), inverta: aqui é (x, y).",
        "Vários no mesmo tile: a célula do mapa empilha t1+t5; esta tabela lista cada um.",
        "★ no mapa = Query estes primeiro. Walkers ANDAM — carregue ACHEA23, pause, clique já.",
    ]
    for i, line in enumerate(notes, start=2):
        lg.cell(i, 1, line)
        lg.merge_cells(start_row=i, start_column=1, end_row=i, end_column=12)

    query_key = {(q["slot"], q["type"], q["x"], q["y"]) for q in query}
    hdr_row = 8
    headers = [
        "slot",
        "tipo",
        "nome",
        "x",
        "y",
        "Excel",
        "estado",
        "facing",
        "no tile",
        "perto",
        "casa_x",
        "casa_y",
        "casa",
        "pilha",
        "Query?",
        "Nome",
        "Frase",
    ]
    for c, h in enumerate(headers, start=1):
        cell = lg.cell(hdr_row, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    type_fills = {t: PatternFill("solid", fgColor=c) for t, c in TYPE_FILL.items()}
    body = Font(name="Calibri", size=9)
    for i, row in enumerate(rows):
        r = hdr_row + 1 + i
        hx, hy = row["home"] if row["home"] else ("", "")
        is_q = (row["slot"], row["type"], row["x"], row["y"]) in query_key
        hit = QUERY_HITS.get(row["slot"])
        qmark = "★ PRIMEIRO" if is_q else ("Query'd" if hit else "")
        person, quote = hit if hit else ("", "")
        vals = [
            row["slot"],
            row["type"],
            TYPE_LABEL[row["type"]],
            row["x"],
            row["y"],
            row["excel"],
            row["state"],
            row["face"],
            row["on"],
            ", ".join(row["near"][:4]),
            hx,
            hy,
            row["home_name"],
            row["stack"],
            qmark,
            person,
            quote,
        ]
        fill = THICK_YELLOW if is_q else type_fills[row["type"]]
        for c, v in enumerate(vals, start=1):
            cell = lg.cell(r, c, v if v != "" else None)
            cell.font = body
            cell.fill = fill
    lg.auto_filter.ref = f"A{hdr_row}:Q{hdr_row + len(rows)}"
    lg.freeze_panes = f"A{hdr_row + 1}"
    widths = [7, 7, 14, 5, 5, 8, 8, 8, 14, 42, 8, 8, 14, 7, 14, 18, 28]
    for i, w in enumerate(widths, start=1):
        lg.column_dimensions[get_column_letter(i)].width = w


def write_query(wb: Workbook, query: list[dict], rows: list[dict]) -> None:
    qs = wb.create_sheet("query", 0)
    qs["A1"] = "Query estes primeiro"
    qs["A1"].font = Font(name="Calibri", size=14, bold=True)
    howto = [
        "1. Abra ACHEA23.SAV no DosBox (pasta Achea.sav). Pause o jogo.",
        "2. Vá no tile (x, y) abaixo — mesma grade do Achea_grid_v3. Clique na PESSOA, não no prédio.",
        "3. Manda o texto do Query + um print. Um tipo por vez basta.",
        "Walkers andam. Se a pessoa não estiver no tile, olha a rua/plaza do lado (1–2 tiles).",
        "Tipos 3 (bárbaro) e 7 (revoltoso) não existem nesta save — não tem o que clicar.",
    ]
    for i, line in enumerate(howto, start=2):
        qs.cell(i, 1, line)
        qs.merge_cells(start_row=i, start_column=1, end_row=i, end_column=10)

    counts: dict[int, int] = defaultdict(int)
    for row in rows:
        counts[row["type"]] += 1

    qs.cell(8, 1, "resumo desta save")
    qs.cell(8, 1).font = Font(name="Calibri", size=11, bold=True)
    qs.cell(9, 1, f"{len(rows)} walkers vivos em {len({(r['x'], r['y']) for r in rows})} tiles")
    qs.merge_cells("A9:F9")
    r = 10
    qs.cell(r, 1, "tipo")
    qs.cell(r, 2, "qtd")
    qs.cell(r, 3, "nome")
    qs.cell(r, 4, "exemplo 1 (x,y)")
    qs.cell(r, 5, "exemplo 2 (x,y)")
    qs.cell(r, 6, "hint (não é o nome do Query)")
    for c in range(1, 7):
        qs.cell(r, c).font = HEADER_FONT
        qs.cell(r, c).fill = HEADER_FILL
    by_type: dict[int, list[dict]] = defaultdict(list)
    for q in query:
        by_type[q["type"]].append(q)
    type_fills = {t: PatternFill("solid", fgColor=c) for t, c in TYPE_FILL.items()}
    body = Font(name="Calibri", size=10)
    for i, typ in enumerate(sorted(set(counts) | {3, 7})):
        rr = 11 + i
        examples = by_type.get(typ, [])
        e1 = f"({examples[0]['x']},{examples[0]['y']})" if examples else "—"
        e2 = f"({examples[1]['x']},{examples[1]['y']})" if len(examples) > 1 else "—"
        vals = [typ, counts.get(typ, 0), TYPE_LABEL[typ], e1, e2, TYPE_HINT[typ]]
        fill = type_fills.get(typ, PatternFill("solid", fgColor="DDDDDD"))
        for c, v in enumerate(vals, start=1):
            cell = qs.cell(rr, c, v)
            cell.font = body
            cell.fill = fill

    hdr = 19
    qs.cell(hdr, 1, "células para clicar (1–2 por tipo; ★ = os 5 do resumo)")
    qs.cell(hdr, 1).font = Font(name="Calibri", size=11, bold=True)
    qs.merge_cells(start_row=hdr, start_column=1, end_row=hdr, end_column=10)
    headers = [
        "#",
        "tipo",
        "x",
        "y",
        "Excel",
        "slot",
        "estado",
        "no tile",
        "perto",
        "como achar",
    ]
    for c, h in enumerate(headers, start=1):
        cell = qs.cell(hdr + 1, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    # First of each type = the 5 to Query first; second = backup.
    first_of_type: dict[int, dict] = {}
    for q in query:
        first_of_type.setdefault(q["type"], q)
    star_n = 0
    alt_n = 0
    for q in query:
        is_first = first_of_type[q["type"]] is q
        if is_first:
            star_n += 1
            tag = f"★{star_n}"
        else:
            alt_n += 1
            tag = f"alt {alt_n}"
        rr = hdr + 1 + star_n + alt_n
        near = ", ".join(q["near"][:3]) or q["on"]
        how = f"{q['on']} em ({q['x']},{q['y']}) — {near}"
        if not is_first:
            how = f"reserva t{q['type']} — clique se o ★ já andou. {how}"
        vals = [
            tag,
            q["type"],
            q["x"],
            q["y"],
            q["excel"],
            q["slot"],
            q["state"],
            q["on"],
            ", ".join(q["near"][:3]),
            how,
        ]
        fill = THICK_YELLOW if is_first else type_fills[q["type"]]
        for c, v in enumerate(vals, start=1):
            cell = qs.cell(rr, c, v)
            cell.font = Font(name="Calibri", size=10, bold=is_first)
            cell.fill = fill
            cell.border = QUERY_BORDER if is_first else THIN

    widths = [8, 7, 5, 5, 8, 7, 8, 14, 36, 62]
    for i, w in enumerate(widths, start=1):
        qs.column_dimensions[get_column_letter(i)].width = w
    qs.freeze_panes = "A21"


def main() -> None:
    if not SAV.is_file():
        raise SystemExit(f"missing SAV: {SAV}")
    game = SAV.parent.parent
    city = load_city_from_sav(SAV)
    walkers = live_walkers(load_walkers_from_sav(SAV, game=game))
    rows = build_rows(city, walkers)
    query = pick_query_examples(rows, per_type=2)

    wb = Workbook()
    write_map(wb, city, rows, query)
    write_legend(wb, rows, query)
    write_query(wb, query, rows)

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"=== WROTE {OUT_XLSX}  {OUT_XLSX.stat().st_size} B ===")
    print(f"live walkers: {len(rows)}")
    counts = defaultdict(int)
    for row in rows:
        counts[row["type"]] += 1
    print("types:", dict(sorted(counts.items())))
    stacks = sum(1 for row in rows if row["stack"] > 1)
    print(f"stacked records: {stacks}  unique tiles: {len({(r['x'], r['y']) for r in rows})}")
    print("Query first:")
    seen = set()
    for q in query:
        mark = "*" if q["type"] not in seen else " "
        seen.add(q["type"])
        print(
            f"  {mark} t{q['type']} ({q['x']},{q['y']}) Excel {q['excel']} "
            f"W{q['slot']} on {q['on']} near {q['near'][:3]}"
        )


if __name__ == "__main__":
    main()
