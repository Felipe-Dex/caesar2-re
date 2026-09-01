#!/usr/bin/env python3
"""Compare the user's by-eye Achea.xlsx to ACHEA23, then export a numbered grid.

Does not copy the SAV. Does not copy the Downloads xlsx into git.
"""

from __future__ import annotations

import sys
from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from app.city_map import (  # noqa: E402
    FLAG_PAD,
    FLAG_RIVER,
    ID_TERRAIN_MAX,
    MAP_H,
    MAP_W,
    load_city_from_sav,
)

SAV = Path(r"C:\Users\Felip\OneDrive\Games\Caesar2\Achea.sav\ACHEA23.SAV")
USER_XLSX = Path(r"C:\Users\Felip\Downloads\Achea.xlsx")
OUT_XLSX = ROOT / "findings" / "Achea_grid.xlsx"
OUT_XLSX_V3 = ROOT / "findings" / "Achea_grid_v3.xlsx"
OUT_CSV = ROOT / "findings" / "Achea_grid.csv"
OUT_LEGEND_CSV = ROOT / "findings" / "Achea_grid_legend.csv"

# 0xFA +19 lo-nibble = goods (ghidra_tile.md). User named two on Achea.
# 0xFA origin +19 lo-nibble. UI names from Achea + 20230610 (findings/factory.md).
FA_GOODS = {
    0: "Bakery",
    1: "Winery",
    2: "Butcher",
    5: "Lead Works",
    11: "Stone Works",
    14: "Ivory Dealer",
}

# Confirmed / range names from findings (achea.md, build_palette.md) + grid v3 dump.
# N is not renumbered: new names go in KNOWN *and* PROMOTED.
KNOWN: dict[int, str] = {}
KNOWN[0x78] = "Garden"
KNOWN[0x79] = "Garden"
KNOWN[0x7A] = "Garden"
KNOWN[0x7B] = "Garden"
KNOWN[0x7C] = "Plaza 1"
KNOWN[0x7D] = "Plaza"
KNOWN[0x7E] = "Plaza est"
KNOWN[0x82] = "Tent"
for _i in range(0x83, 0xA2):
    KNOWN[_i] = "Casa"
KNOWN[0xA2] = "Shrine 1"
KNOWN[0xA3] = "Shrine 2"
KNOWN[0xA4] = "Shrine 3"
KNOWN[0xA5] = "Shrine 4"
KNOWN[0xA6] = "Temple"
KNOWN[0xA7] = "Temple"
KNOWN[0xA8] = "Temple"
KNOWN[0xAA] = "Basilica"
KNOWN[0xAB] = "Basilica"
KNOWN[0xAC] = "Basilica 4"
KNOWN[0xAF] = "Aventine"
KNOWN[0xB2] = "Janiculan 1"
KNOWN[0xB3] = "Janiculan 2"
KNOWN[0xB4] = "Janiculan 4"
KNOWN[0xB7] = "Palatine"
KNOWN[0xBE] = "Reservatorio"
KNOWN[0xBF] = "Tower"
KNOWN[0xC0] = "Gate"
KNOWN[0xC1] = "Wall N-S?"
KNOWN[0xC2] = "Wall"
for _i in range(0xCF, 0xD7):
    KNOWN[_i] = "Aqueduto"
KNOWN[0xDC] = "Fountain 2"
KNOWN[0xDD] = "Fountain 1"
KNOWN[0xDE] = "Fountain 4"
KNOWN[0xDF] = "Baths 1"
KNOWN[0xE0] = "Baths"
KNOWN[0xE2] = "Baths 4"
KNOWN[0xE3] = "Praefecture"
KNOWN[0xE4] = "Barracks"
KNOWN[0xD7] = "Well"
KNOWN[0xE5] = "Theater"
KNOWN[0xE6] = "Odeum"
KNOWN[0xE8] = "Colosseum"
KNOWN[0xE9] = "Circus"
KNOWN[0xEA] = "Circus"
KNOWN[0xEB] = "Circus"
KNOWN[0xEC] = "Circus"
KNOWN[0xED] = "C.Maximus"
KNOWN[0xEE] = "C.Maximus"
KNOWN[0xF3] = "Grammaticus"
KNOWN[0xF4] = "Rhetor"
KNOWN[0xF5] = "Library"
KNOWN[0xFA] = "Factory"
KNOWN[0xFB] = "Hospital"
KNOWN[0xFD] = "Market 2"
KNOWN[0xFE] = "Market 3"
KNOWN[0xFF] = "Market 4"

# Had a Desconhecido N, then named — stay in the N sequence.
PROMOTED = {
    0x78, 0x79, 0x7A, 0x7B,
    0xA2, 0xA3, 0xA4, 0xA5,
    0xAF, 0xB2, 0xB3, 0xB4, 0xB7,
    0xDC, 0xDD, 0xDE, 0xDF,
    0xE0, 0xE2, 0xE3, 0xE4, 0xE6,
    0xE9, 0xEA, 0xEB, 0xEC,
    0xF3, 0xF4, 0xF5, 0xFA, 0xFB,
    0xFD, 0xFE, 0xFF,
}
CORE_KNOWN = {k for k in KNOWN if k not in PROMOTED}

# User-sheet synonyms → our known name (for alignment scoring).
USER_ALIAS = {
    "tower": "Tower",
    "wall": "Wall",
    "gate": "Gate",
    "gate?": "Gate",
    "road": "Road",
    "circus": "Circus",
    "c.maximus": "C.Maximus",
    "colosseum": "Colosseum",
    "basilica": "Basilica",
    "janiculan": "Janiculan 2",
    "aventine": "Aventine",
    "palatine": "Palatine",
    "reservour": "Reservatorio",
    "reservoir": "Reservatorio",
    "plaza": "Plaza 1",
    "baths": "Baths",
    "market": "Market 2",
    "temple": "Temple",
    "barracks": "Barracks",
    "hospital": "Hospital",
    "library": "Library",
    "praefacture": "Praefecture",
    "praefecture": "Praefecture",
    "bridge": None,
}


CLASS_OF_NAME = {
    "Tent": "house",
    "Casa": "house",
    "Temple": "worship",
    "Shrine 1": "worship",
    "Shrine 2": "worship",
    "Shrine 3": "worship",
    "Shrine 4": "worship",
    "Basilica": "worship",
    "Basilica 4": "worship",
    "Aventine": "forum",
    "Janiculan 1": "forum",
    "Janiculan 2": "forum",
    "Janiculan 4": "forum",
    "Palatine": "forum",
    "Reservatorio": "water",
    "Aqueduto": "water",
    "Fountain 1": "water",
    "Fountain 2": "water",
    "Fountain 4": "water",
    "Rio": "water",
    "Tower": "wall",
    "Gate": "wall",
    "Wall": "wall",
    "Wall N-S?": "wall",
    "Road": "road",
    "Garden": "garden",
    "Plaza 1": "plaza",
    "Plaza": "plaza",
    "Plaza est": "plaza",
    "C.Maximus": "ent",
    "Colosseum": "ent",
    "Circus": "ent",
    "Odeum": "ent",
    "Market 1": "market",
    "Market 2": "market",
    "Market 3": "market",
    "Market 4": "market",
    "Baths": "amenity",
    "Baths 1": "amenity",
    "Baths 3": "amenity",
    "Baths 4": "amenity",
    "Hospital": "amenity",
    "Grammaticus": "edu",
    "Rhetor": "edu",
    "Library": "edu",
    "Praefecture": "security",
    "Barracks": "security",
    "Factory": "industry",
    "Bakery": "industry",
    "Winery": "industry",
    "Butcher": "industry",
    "Lead Works": "industry",
    "Stone Works": "industry",
    "Ivory Dealer": "industry",
}

FILLS = {
    "house": "F4A6A6",
    "forum": "C5B0D5",
    "water": "6BAED6",
    "wall": "C4A574",
    "road": "B0B0B0",
    "plaza": "D5C4A1",
    "garden": "82E0AA",
    "ent": "FDAE6B",
    "worship": "E8C547",
    "market": "A1D99B",
    "amenity": "76D7C4",
    "edu": "85C1E9",
    "security": "F5B7B1",
    "industry": "D4A017",
    "unknown": "FFF2A8",
    "terrain": "F7F7F7",
    "header": "1F4E79",
}

THIN = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


def same_id_blobs(ids: list[list[int]]) -> list[dict]:
    seen = [[False] * MAP_W for _ in range(MAP_H)]
    out: list[dict] = []
    for y in range(MAP_H):
        for x in range(MAP_W):
            tid = ids[y][x]
            if tid < ID_TERRAIN_MAX or seen[y][x]:
                continue
            stack = [(x, y)]
            seen[y][x] = True
            cells: list[tuple[int, int]] = []
            while stack:
                cx, cy = stack.pop()
                cells.append((cx, cy))
                for dx, dy in ((1, 0), (-1, 0), (0, 1), (0, -1)):
                    nx, ny = cx + dx, cy + dy
                    if (
                        0 <= nx < MAP_W
                        and 0 <= ny < MAP_H
                        and not seen[ny][nx]
                        and ids[ny][nx] == tid
                    ):
                        seen[ny][nx] = True
                        stack.append((nx, ny))
            xs = [c[0] for c in cells]
            ys = [c[1] for c in cells]
            out.append(
                {
                    "id": tid,
                    "n": len(cells),
                    "xmin": min(xs),
                    "xmax": max(xs),
                    "ymin": min(ys),
                    "ymax": max(ys),
                    "w": max(xs) - min(xs) + 1,
                    "h": max(ys) - min(ys) + 1,
                    "cells": cells,
                }
            )
    return out


def load_user_grid() -> dict[tuple[int, int], str]:
    wb = load_workbook(USER_XLSX, data_only=True)
    ws = wb.active
    grid: dict[tuple[int, int], str] = {}
    # Row 1 = x headers 55..79 in cols 2..26. Col A = y headers.
    # Rows 24–26 have no y header; they continue 22, 23, 24.
    last_y = None
    for r in range(2, ws.max_row + 1):
        yh = ws.cell(r, 1).value
        if yh is None or yh == "":
            if last_y is None:
                continue
            y = last_y + 1
        else:
            y = int(yh)
        last_y = y
        for c in range(2, ws.max_column + 1):
            xh = ws.cell(1, c).value
            if xh is None:
                continue
            x = int(xh)
            v = ws.cell(r, c).value
            if v is None or str(v).strip() == "":
                continue
            grid[(x, y)] = str(v).strip()
    return grid


def describe_user_sheet() -> None:
    wb = load_workbook(USER_XLSX, data_only=True)
    print("=== USER XLSX ===")
    print(f"  path: {USER_XLSX}")
    print(f"  size: {USER_XLSX.stat().st_size} B")
    print(f"  sheets: {wb.sheetnames}")
    ws = wb.active
    print(f"  active: {ws.title}  {ws.max_row}x{ws.max_column}  freeze={ws.freeze_panes}")
    xs = [ws.cell(1, c).value for c in range(2, ws.max_column + 1)]
    ys = [ws.cell(r, 1).value for r in range(2, ws.max_row + 1)]
    print(f"  col headers (row1): {xs}")
    print(f"  row headers (colA): {ys}")
    print("  NOT 80x80 — window only.")


def main() -> None:
    user: dict[tuple[int, int], str] = {}
    if USER_XLSX.is_file():
        describe_user_sheet()
        user = load_user_grid()
        print(f"  labeled cells: {len(user)}")
        print(f"  user bbox x={min(x for x,_ in user)}..{max(x for x,_ in user)} "
              f"y={min(y for _,y in user)}..{max(y for _,y in user)}")
        print(f"  labels: {sorted(set(user.values()))}")
    else:
        print(f"=== USER XLSX missing ({USER_XLSX}) — skip alignment ===")

    city = load_city_from_sav(SAV)
    ids = [[city.tile(x, y).terrain_id for x in range(MAP_W)] for y in range(MAP_H)]
    flags = [[city.tile(x, y).flags for x in range(MAP_W)] for y in range(MAP_H)]
    specials = [[city.tile(x, y).special for x in range(MAP_W)] for y in range(MAP_H)]
    blobs = same_id_blobs(ids)

    # 0xFA industry: +19 lo-nibble on the home tile (ghidra_tile.md).
    fa_name_at: dict[tuple[int, int], str] = {}
    for b in blobs:
        if b["id"] != 0xFA:
            continue
        nibs = {specials[cy][cx] & 0x0F for cx, cy in b["cells"] if specials[cy][cx] & 0x0F}
        goods = next(iter(nibs), 0) if len(nibs) == 1 else (max(nibs) if nibs else 0)
        name = FA_GOODS.get(goods, "Factory")
        for cx, cy in b["cells"]:
            fa_name_at[(cx, cy)] = name

    # Assign Desconhecido N to blobs not in CORE_KNOWN (same-id 4-conn).
    # Order: north-then-west origin. Promoted names (now in KNOWN) keep their
    # old N off the legend so remaining N do not shift.
    numbered_blobs = [b for b in blobs if b["id"] not in CORE_KNOWN]
    numbered_blobs.sort(key=lambda b: (b["ymin"], b["xmin"], b["id"]))
    label_at: dict[tuple[int, int], str] = {}
    unk_meta: list[dict] = []
    for n, b in enumerate(numbered_blobs, start=1):
        name = f"Desconhecido {n}"
        for x, y in b["cells"]:
            label_at[(x, y)] = name
        unk_meta.append({**b, "n_label": n, "name": name})
    still_unknown = [b for b in unk_meta if b["id"] not in KNOWN]

    # Named + river overlay (building name wins). Terrain + FLAG_PAD = Road.
    grid_text = [[""] * MAP_W for _ in range(MAP_H)]
    grid_class = [["terrain"] * MAP_W for _ in range(MAP_H)]
    for y in range(MAP_H):
        for x in range(MAP_W):
            tid = ids[y][x]
            if tid == 0xFA:
                grid_text[y][x] = fa_name_at[(x, y)]
                grid_class[y][x] = CLASS_OF_NAME[grid_text[y][x]]
            elif tid in KNOWN:
                grid_text[y][x] = KNOWN[tid]
                grid_class[y][x] = CLASS_OF_NAME[KNOWN[tid]]
            elif tid >= ID_TERRAIN_MAX:
                grid_text[y][x] = label_at[(x, y)]
                grid_class[y][x] = "unknown"
            elif flags[y][x] & FLAG_PAD:
                grid_text[y][x] = "Road"
                grid_class[y][x] = "road"
            elif flags[y][x] & FLAG_RIVER:
                grid_text[y][x] = "Rio"
                grid_class[y][x] = "water"

    # --- alignment vs user sheet ---
    print("\n=== ALIGNMENT (user window vs chunk 13) ===")
    agree = 0
    disagree = 0
    user_empty_we_have = 0
    they_named_terrain = 0
    misses: list[str] = []
    by_user_label: dict[str, Counter] = {}
    for (x, y), ulab in sorted(user.items()):
        our = grid_text[y][x]
        tid = ids[y][x]
        key = ulab.lower()
        by_user_label.setdefault(ulab, Counter())[f"0x{tid:02X}" if tid >= 0x78 else f"terra/0x{tid:02X}"] += 1
        alias = USER_ALIAS.get(key, None)
        if our == "":
            they_named_terrain += 1
            disagree += 1
            if len(misses) < 40:
                misses.append(f"  ({x},{y}) user={ulab!r} ours=TERRAIN id=0x{tid:02X}")
            continue
        if alias is None:
            # they guessed a name we do not close — not an alignment error if
            # they landed on a building blob.
            if tid >= ID_TERRAIN_MAX:
                agree += 1  # same cell is a building; name is their guess
            else:
                disagree += 1
            continue
        # alias is our known family
        if alias == our or our.startswith(alias.split()[0]):
            agree += 1
        elif alias == "Wall" and our in ("Wall", "Wall N-S?", "Tower", "Gate"):
            agree += 1
        elif alias == "Circus" and our in ("Circus", "C.Maximus"):
            agree += 1  # old by-eye sheet used Circus for the long oval
        elif alias == "C.Maximus" and our != "C.Maximus":
            disagree += 1
            misses.append(f"  ({x},{y}) user={ulab!r} ours={our!r} id=0x{tid:02X}  << C.Maximus misplaced")
        else:
            disagree += 1
            if len(misses) < 50:
                misses.append(f"  ({x},{y}) user={ulab!r} ours={our!r} id=0x{tid:02X}")

    # cells we have named in their window that they left blank
    ux0, ux1 = 55, 79
    uy0, uy1 = 0, 24
    blank_but_ours = []
    for y in range(uy0, uy1 + 1):
        for x in range(ux0, ux1 + 1):
            if (x, y) in user:
                continue
            if grid_text[y][x] and grid_class[y][x] != "terrain":
                if grid_text[y][x] == "Rio":
                    continue
                blank_but_ours.append((x, y, grid_text[y][x], ids[y][x]))
                user_empty_we_have += 1

    print(f"  user labeled {len(user)} cells in ~x=55..79 y=0..24")
    print(f"  name-family agree (closed names): {agree}")
    print(f"  clear mismatches (incl. Circus / terrain stamps): {disagree}")
    print(f"  they labeled terrain: {they_named_terrain}")
    print(f"  they left blank but we have a building: {user_empty_we_have}")
    print("  per-label -> actual tile[+0]:")
    for lab, ctr in sorted(by_user_label.items()):
        print(f"    {lab:14s} {dict(ctr)}")
    print("  sample mismatches:")
    for line in misses[:25]:
        print(line)

    print("  blank-but-ours (first 20):")
    for x, y, name, tid in blank_but_ours[:20]:
        print(f"    ({x},{y}) {name} id=0x{tid:02X}")

    # Landmark check
    print("\n=== LANDMARKS ===")
    checks = [
        ((55, 1), "Tower"),
        ((78, 1), "Tower"),
        ((70, 1), "Gate"),
        ((71, 13), "Colosseum"),
        ((71, 17), "Basilica"),
        ((71, 10), "Basilica"),
        ((71, 6), "Janiculan 2"),
        ((74, 6), "Barracks"),
        ((2, 70), "Aventine"),
        ((71, 25), "C.Maximus"),
        ((71, 20), "Palatine"),
        ((70, 2), "Road"),
        ((70, 45), "Plaza 1"),
        ((70, 44), "Plaza est"),
        ((46, 5), "Grammaticus"),
    ]
    for (x, y), expect in checks:
        ul = user.get((x, y), "(vazio)")
        print(f"  ({x},{y}) ours={grid_text[y][x]!r} id=0x{ids[y][x]:02X}  user={ul!r}  expect~{expect}")

    print(f"\n=== DESCONHECIDO === {len(still_unknown)} blobs (N stable; {len(unk_meta) - len(still_unknown)} promoted)")
    # Prefer city-core big blobs for "first 5"
    core = [b for b in still_unknown if not (b["ymin"] >= 62 or b["id"] in (0xFB, 0xF5))]
    core_by_size = sorted(core, key=lambda b: (-b["n"], b["ymin"], b["xmin"]))
    print("  first 5 (city-ish, by size):")
    for b in core_by_size[:5]:
        print(
            f"    N={b['n_label']} id=0x{b['id']:02X} {b['w']}x{b['h']} "
            f"n={b['n']} bbox=({b['xmin']},{b['ymin']})-({b['xmax']},{b['ymax']})"
        )
    print("  leftover after v3 dump (first 8 by size):")
    for b in core_by_size[:8]:
        print(
            f"    N={b['n_label']} id=0x{b['id']:02X} {b['w']}x{b['h']} "
            f"n={b['n']} bbox=({b['xmin']},{b['ymin']})-({b['xmax']},{b['ymax']})"
        )

    # --- Excel ---
    wb = Workbook()
    ws = wb.active
    ws.title = "mapa"

    header_font = Font(name="Calibri", size=8, bold=True, color="FFFFFF")
    cell_font = Font(name="Calibri", size=7)
    header_fill = PatternFill("solid", fgColor=FILLS["header"])
    header_align = Alignment(horizontal="center", vertical="center")
    cell_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fill_cache = {
        k: PatternFill("solid", fgColor=v) for k, v in FILLS.items() if k != "header"
    }

    ws.cell(1, 1, "y\\x")
    ws.cell(1, 1).font = header_font
    ws.cell(1, 1).fill = header_fill
    ws.cell(1, 1).alignment = header_align
    for x in range(MAP_W):
        cell = ws.cell(1, x + 2, x)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = THIN
    for y in range(MAP_H):
        cell = ws.cell(y + 2, 1, y)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = THIN
        for x in range(MAP_W):
            text = grid_text[y][x]
            cell = ws.cell(y + 2, x + 2, text if text else None)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = THIN
            cls = grid_class[y][x]
            if text:
                cell.fill = fill_cache[cls]
            else:
                cell.fill = fill_cache["terrain"]

    ws.freeze_panes = "B2"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 14
    ws.column_dimensions["A"].width = 4.5
    for x in range(MAP_W):
        ws.column_dimensions[get_column_letter(x + 2)].width = 5.2
    for y in range(MAP_H):
        ws.row_dimensions[y + 2].height = 12
    ws.auto_filter.ref = None
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # Legend
    lg = wb.create_sheet("legenda")
    lg["A1"] = "Como ler"
    lg["A1"].font = Font(name="Calibri", size=12, bold=True)
    notes = [
        "(0,0) = ponta NORTE do losango (não é o canto superior-esquerdo da tela 2D).",
        "Iso: direita = x−y; baixo = x+y. Ponta OESTE = (0,79). Leste = x alto, y baixo.",
        "Números nas bordas são tile x=0..79 e y=0..79 do chunk 13 (mesmo sistema do host).",
        "Célula vazia = terreno sem prédio. Rio = tile[+1] & 0x10. Road = terreno com tile[+1] & 0x20 (BT4:BT7 / BT7:BZ7).",
        "Palatine = 0xB7 4×4 FECHADO (D56). Aventine = 0xAF. Janiculan 1/2/4 = 0xB2/0xB3/0xB4 (3º ausente nesta save).",
        "Basilica = 0xAB; Basilica 4 (mais evoluída) = 0xAC. Temple = 0xA6–0xA8. Shrine 1–4 = 0xA2–0xA5 (2º = walker).",
        "Plaza: 0x7C = nível 1 (BT47); 0x7E = com estátua (BT46); 0x7D = junta. Garden = 0x78–0x7B (era hyp Plaza).",
        "Market 2/3/4 = 0xFD pouco / 0xFE frequente / 0xFF thriving. Fountain 1/2/4 = 0xDD/0xDC/0xDE. Baths 1/4 = 0xDF/0xE2; 0xE0 sem estágio.",
        "Factory 0xFA: Winery (+19=1) / Lead Works (+19=5) / resto Factory. Grammaticus 0xF3, Rhetor 0xF4, Odeum 0xE6.",
        "Desconhecido N = blob 4-conexo do mesmo id ≥0x78 ainda sem nome. N NÃO renumera.",
        "Cores: vermelho=casa, roxo=fórum, dourado=culto, azul=água, marrom=muro, cinza=estrada, areia=plaza, verde-claro=jardim, laranja=entretenimento, verde=mercado, teal=termas, azul-claro=edu, rosa=segurança, ocre=fábrica, amarelo=desconhecido.",
    ]
    for i, line in enumerate(notes, start=2):
        lg.cell(i, 1, line)
        lg.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)

    hdr_row = 15
    headers = ["N", "id_hex", "id_dec", "w", "h", "x0", "y0", "x1", "y1", "count", "nome", "nota"]
    for c, h in enumerate(headers, start=1):
        cell = lg.cell(hdr_row, c, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    unk_yellow = PatternFill("solid", fgColor=FILLS["unknown"])
    # Legend sorted by size (easier to look up in-game), N stays the grid number.
    for i, b in enumerate(sorted(still_unknown, key=lambda z: (-z["n"], z["ymin"], z["xmin"]))):
        r = hdr_row + 1 + i
        vals = [
            b["n_label"],
            f"0x{b['id']:02X}",
            b["id"],
            b["w"],
            b["h"],
            b["xmin"],
            b["ymin"],
            b["xmax"],
            b["ymax"],
            b["n"],
            b["name"],
            "olha no jogo",
        ]
        for c, v in enumerate(vals, start=1):
            cell = lg.cell(r, c, v)
            cell.font = Font(name="Calibri", size=9)
            cell.fill = unk_yellow
    lg.auto_filter.ref = f"A{hdr_row}:L{hdr_row + len(still_unknown)}"
    lg.freeze_panes = f"A{hdr_row + 1}"
    widths = [6, 10, 8, 5, 5, 6, 6, 6, 6, 8, 18, 14]
    for i, w in enumerate(widths, start=1):
        lg.column_dimensions[get_column_letter(i)].width = w

    # Known-names sheet (short)
    kn = wb.create_sheet("nomes")
    kn["A1"] = "id"
    kn["B1"] = "nome"
    kn["C1"] = "conf"
    for c in range(1, 4):
        kn.cell(1, c).font = header_font
        kn.cell(1, c).fill = header_fill
    known_rows = [
        ("0x78–0x7B", "Garden (D42=0x7B, D142=0x78; 0x79/0x7A vizinhos)", "v3 HIGH"),
        ("0x7C", "Plaza 1 (BT47)", "v3 HIGH"),
        ("0x7D", "Plaza (junta)", "v3"),
        ("0x7E", "Plaza estátua (BT46)", "v3 HIGH"),
        ("terreno +0x20", "Road (BT4:BT7 e BT7:BZ7; era vazio)", "v3 HIGH"),
        ("0x82", "Tent", "A/B"),
        ("0x83–0xA1", "Casa (graus de habitação)", "range"),
        ("0xA2", "Shrine 1 (D138)", "§11"),
        ("0xA3", "Shrine 2 (walker perto + §11)", "walker HIGH"),
        ("0xA4", "Shrine 3 (D14, D95)", "v3 HIGH"),
        ("0xA5", "Shrine 4 (D94)", "v3 HIGH"),
        ("0xA6–0xA8", "Temple 2×2", "Q&A HIGH"),
        ("0xAB", "Basilica 3×3", "Q&A HIGH"),
        ("0xAC", "Basilica 4 mais evoluída", "v3 HIGH"),
        ("0xAF", "Aventine 2×2", "Q&A HIGH"),
        ("0xB2", "Janiculan 1 (D61)", "v3 HIGH"),
        ("0xB3", "Janiculan 2 (era D8)", "v3 HIGH"),
        ("0xB4", "Janiculan 4 mais evoluída (D97)", "v3 HIGH"),
        ("0xB7", "Palatine 4×4 (D56) FECHADO", "v3 HIGH"),
        ("0xBE", "Reservatorio", "A/B"),
        ("0xBF", "Tower", "Q&A HIGH"),
        ("0xC0", "Gate", "Q&A HIGH"),
        ("0xC1", "Wall N-S?", "geom"),
        ("0xC2", "Wall E-W", "Q&A HIGH"),
        ("0xCF–0xD6", "Aqueduto", "FELIPE01"),
        ("0xDD", "Fountain 1 (D18)", "v3 HIGH"),
        ("0xDC", "Fountain 2 (D52)", "v3 HIGH"),
        ("0xDE", "Fountain 4 (D83)", "v3 HIGH"),
        ("0xDF", "Baths 1 (D51)", "v3 HIGH"),
        ("0xE0", "Baths (estágio não dito)", "Q&A HIGH"),
        ("0xE2", "Baths 4 (D82)", "v3 HIGH"),
        ("0xE3", "Praefecture", "Q&A HIGH"),
        ("0xE4", "Barracks 3×3", "Q&A HIGH"),
        ("0xE6", "Odeum (D41)", "v3 HIGH"),
        ("0xE8", "Colosseum", "Q&A HIGH"),
        ("0xE9 + 0xEA", "Circus 3×6 (35,38); walker 0xEA", "§11 + walker"),
        ("0xEB + 0xEC", "Circus 3×6 (62,2)", "Q&A HIGH"),
        ("0xED + 0xEE", "C.Maximus 4×8", "Q&A HIGH"),
        ("0xF3", "Grammaticus (D5, D125, D147)", "v3 HIGH"),
        ("0xF4", "Rhetor (D84)", "v3 HIGH"),
        ("0xF5", "Library 3×3", "Q&A HIGH"),
        ("0xFA +19=1", "Winery (D22 norte, perto D13)", "v3 HIGH"),
        ("0xFA +19=5", "Lead Works (D40, D22 sul)", "v3 HIGH"),
        ("0xFA outro", "Factory", "v3"),
        ("0xFB", "Hospital 3×3", "Q&A HIGH"),
        ("0xFD", "Market 2 pouco uso", "v3 HIGH"),
        ("0xFE", "Market 3 uso frequente (D10)", "v3 HIGH"),
        ("0xFF", "Market 4 thriving (D46)", "v3 HIGH"),
        ("tile[+1] & 0x10", "Rio", "flag"),
    ]
    for i, row in enumerate(known_rows, start=2):
        for c, v in enumerate(row, start=1):
            kn.cell(i, c, v)
    kn.column_dimensions["A"].width = 18
    kn.column_dimensions["B"].width = 52
    kn.column_dimensions["C"].width = 16
    kn.freeze_panes = "A2"

    # Progress: named vs unknown building tiles / unique ids
    bld_ids = Counter(tid for row in ids for tid in row if tid >= ID_TERRAIN_MAX)
    named_id_set = {i for i in bld_ids if i in KNOWN}
    unk_id_set = {i for i in bld_ids if i not in KNOWN}
    named_tiles = sum(bld_ids[i] for i in named_id_set)
    unk_tiles = sum(bld_ids[i] for i in unk_id_set)
    name_tiles = Counter()
    for y in range(MAP_H):
        for x in range(MAP_W):
            if ids[y][x] >= ID_TERRAIN_MAX:
                name_tiles[grid_text[y][x]] += 1

    pg = wb.create_sheet("progress")
    pg["A1"] = "Achea_grid v3 — progresso"
    pg["A1"].font = Font(name="Calibri", size=12, bold=True)
    summary = [
        ("tiles prédio (id≥0x78)", sum(bld_ids.values())),
        ("tiles nomeados", named_tiles),
        ("tiles Desconhecido", unk_tiles),
        ("% tiles nomeados", f"{100.0 * named_tiles / sum(bld_ids.values()):.1f}%"),
        ("ids únicos prédio", len(bld_ids)),
        ("ids únicos nomeados", len(named_id_set)),
        ("ids únicos Desconhecido", len(unk_id_set)),
        ("blobs Desconhecido (N estável)", len(still_unknown)),
        ("Road (terreno +0x20)", sum(1 for y in range(MAP_H) for x in range(MAP_W) if grid_text[y][x] == "Road")),
    ]
    pg["A3"] = "métrica"
    pg["B3"] = "valor"
    for c in range(1, 3):
        pg.cell(3, c).font = header_font
        pg.cell(3, c).fill = header_fill
    for i, (lab, val) in enumerate(summary, start=4):
        pg.cell(i, 1, lab)
        pg.cell(i, 2, val)

    pg.cell(14, 1, "id Desconhecido")
    pg.cell(14, 2, "tiles")
    pg.cell(14, 3, "hex")
    for c in range(1, 4):
        pg.cell(14, c).font = header_font
        pg.cell(14, c).fill = header_fill
    for i, tid in enumerate(sorted(unk_id_set), start=15):
        pg.cell(i, 1, tid)
        pg.cell(i, 2, bld_ids[tid])
        pg.cell(i, 3, f"0x{tid:02X}")

    pg.cell(14, 5, "nome no mapa")
    pg.cell(14, 6, "tiles")
    pg.cell(14, 5).font = header_font
    pg.cell(14, 5).fill = header_fill
    pg.cell(14, 6).font = header_font
    pg.cell(14, 6).fill = header_fill
    for i, (nm, n) in enumerate(name_tiles.most_common(), start=15):
        pg.cell(i, 5, nm)
        pg.cell(i, 6, n)
    pg.column_dimensions["A"].width = 32
    pg.column_dimensions["B"].width = 14
    pg.column_dimensions["C"].width = 10
    pg.column_dimensions["E"].width = 22
    pg.column_dimensions["F"].width = 10
    pg.freeze_panes = "A15"

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX_V3)
    print(f"\n=== WROTE {OUT_XLSX_V3}  {OUT_XLSX_V3.stat().st_size} B ===")
    try:
        wb.save(OUT_XLSX)
        print(f"=== WROTE {OUT_XLSX}  {OUT_XLSX.stat().st_size} B ===")
    except PermissionError:
        alt = OUT_XLSX.with_name("Achea_grid_new.xlsx")
        print(f"LOCKED {OUT_XLSX} — v3 already written; skip {alt}")
    print(
        f"progress: named {named_tiles}/{sum(bld_ids.values())} tiles  "
        f"ids {len(named_id_set)}/{len(bld_ids)}  unknown blobs={len(still_unknown)}"
    )

    # CSV of the grid (names only)
    with OUT_CSV.open("w", encoding="utf-8", newline="") as f:
        # header
        f.write("y\\x," + ",".join(str(x) for x in range(MAP_W)) + "\n")
        for y in range(MAP_H):
            f.write(str(y) + "," + ",".join(grid_text[y]) + "\n")
    print(f"=== WROTE {OUT_CSV}  {OUT_CSV.stat().st_size} B ===")

    with OUT_LEGEND_CSV.open("w", encoding="utf-8", newline="") as f:
        f.write("N,id_hex,w,h,x0,y0,x1,y1,count,nome,nota\n")
        for b in sorted(still_unknown, key=lambda z: (-z["n"], z["ymin"], z["xmin"])):
            f.write(
                f"{b['n_label']},0x{b['id']:02X},{b['w']},{b['h']},"
                f"{b['xmin']},{b['ymin']},{b['xmax']},{b['ymax']},"
                f"{b['n']},{b['name']},olha no jogo\n"
            )
    print(f"=== WROTE {OUT_LEGEND_CSV}  {OUT_LEGEND_CSV.stat().st_size} B ===")
    print(f"unknown blobs: {len(still_unknown)} (N not renumbered)")


if __name__ == "__main__":
    main()
